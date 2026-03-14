#!/usr/bin/env python3
"""Generate 02_trending_output.xlsx from 02_trending.csv."""

import json
import numpy as np
import pandas as pd
import openpyxl
from pathlib import Path

INPUT_CSV  = Path("/Users/lucaschen/skillgraph/benchmark/test_inputs/02_trending.csv")
OUTPUT_XLS = Path("/Users/lucaschen/skillgraph/outputs/v0/02_trending_output.xlsx")

QUIZ_WEIGHT = 0.05           # per quiz
TEST_WEIGHT = 80 / 6 / 100  # per test ≈ 0.13333...

GRADE_THRESHOLDS = [
    (96.5, "A+"),
    (92.5, "A"),
    (89.5, "A-"),
    (86.5, "B+"),
    (82.5, "B"),
    (79.5, "B-"),
    (76.5, "C+"),
    (72.5, "C"),
    (69.5, "C-"),
    (66.5, "D+"),
    (62.5, "D"),
    (59.5, "D-"),
    (0,    "F"),
]

ALL_GRADES = ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]


def letter_grade(avg):
    for threshold, grade in GRADE_THRESHOLDS:
        if avg >= threshold:
            return grade
    return "F"


def ols_slope(scores):
    x = np.arange(1, len(scores) + 1, dtype=float)
    y = np.array(scores, dtype=float)
    n = len(x)
    return float((n * np.dot(x, y) - x.sum() * y.sum()) / (n * np.dot(x, x) - x.sum() ** 2))


def main():
    df = pd.read_csv(INPUT_CSV)
    quiz_cols = [c for c in df.columns if c.startswith("quiz_")]
    test_cols = [c for c in df.columns if c.startswith("test_")]
    all_score_cols = quiz_cols + test_cols
    n_students = len(df)

    # ── Pass 1: per-student stats ─────────────────────────────────────────────
    student_was = {}
    per_student = {}

    for _, row in df.iterrows():
        sid = str(row["student_id"])
        scores   = [float(row[c]) for c in all_score_cols]
        q_scores = [float(row[c]) for c in quiz_cols]
        t_scores = [float(row[c]) for c in test_cols]

        simple_avg   = round(float(np.mean(scores)), 6)
        weighted_avg = round(
            sum(q * QUIZ_WEIGHT for q in q_scores) + sum(t * TEST_WEIGHT for t in t_scores),
            6,
        )
        slope   = round(ols_slope(scores), 6)
        std_dev = round(float(np.std(scores, ddof=1)), 6)

        student_was[sid] = weighted_avg
        per_student[sid] = {
            "raw_scores":      scores,
            "quiz_scores":     q_scores,
            "test_scores":     t_scores,
            "simple_average":  simple_avg,
            "weighted_average": weighted_avg,
            "slope":           slope,
            "std_dev":         std_dev,
            "highest_score":   float(max(scores)),
            "lowest_score":    float(min(scores)),
        }

    # ── Class-level ───────────────────────────────────────────────────────────
    all_was = list(student_was.values())
    class_wa_mean = float(np.mean(all_was))
    class_wa_std  = float(np.std(all_was, ddof=1))

    # ── Rank (competition rank) ────────────────────────────────────────────────
    sorted_avgs = sorted(set(all_was), reverse=True)
    rank_map = {avg: i + 1 for i, avg in enumerate(sorted_avgs)}
    ranks = {sid: rank_map[wa] for sid, wa in student_was.items()}

    # ── Percentile ────────────────────────────────────────────────────────────
    def compute_percentile(wa):
        count_below = sum(1 for w in all_was if w < wa)
        if n_students <= 1:
            return 0.0
        return round(count_below / (n_students - 1) * 100, 1)

    # ── Z-score ───────────────────────────────────────────────────────────────
    def compute_z_score(wa):
        if class_wa_std == 0:
            return 0.0
        return round((wa - class_wa_mean) / class_wa_std, 2)

    # ── Finalise per-student ──────────────────────────────────────────────────
    for sid in per_student:
        wa = student_was[sid]
        per_student[sid]["letter_grade"]       = letter_grade(wa)
        per_student[sid]["rank"]               = ranks[sid]
        per_student[sid]["percentile"]         = compute_percentile(wa)
        per_student[sid]["z_score"]            = compute_z_score(wa)
        per_student[sid]["above_class_average"] = bool(wa > class_wa_mean)

    # ── Per-assessment stats ──────────────────────────────────────────────────
    per_assessment = {}
    for col in all_score_cols:
        vals     = df[col].astype(float).values
        col_mean = float(np.mean(vals))
        col_std  = float(np.std(vals, ddof=1))

        per_student_z = {}
        for _, row in df.iterrows():
            sid = str(row["student_id"])
            v   = float(row[col])
            z   = 0.0 if col_std == 0 else round((v - col_mean) / col_std, 2)
            per_student_z[sid] = z

        per_assessment[col] = {
            "class_average":        round(col_mean, 6),
            "std_dev":              round(col_std, 6),
            "highest_score":        float(np.max(vals)),
            "lowest_score":         float(np.min(vals)),
            "per_student_z_scores": per_student_z,
        }

    # ── Overall ───────────────────────────────────────────────────────────────
    grade_dist = {g: 0 for g in ALL_GRADES}
    for wa in all_was:
        grade_dist[letter_grade(wa)] += 1

    most_improved_sid = max(
        per_student,
        key=lambda s: (per_student[s]["slope"], student_was[s]),
    )
    most_consistent_sid = min(
        per_student,
        key=lambda s: (per_student[s]["std_dev"], -student_was[s]),
    )

    overall = {
        "class_weighted_average": round(class_wa_mean, 6),
        "class_median":           round(float(np.median(all_was)), 6),
        "grade_distribution":     grade_dist,
        "most_improved_student":  most_improved_sid,
        "most_consistent_student": most_consistent_sid,
    }

    # ── Write Excel ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: Student Stats
    ws1 = wb.active
    ws1.title = "Student Stats"
    ss_headers = (
        ["student_id"]
        + quiz_cols
        + test_cols
        + ["simple_average", "weighted_average", "letter_grade", "slope",
           "std_dev", "highest_score", "lowest_score", "rank",
           "percentile", "z_score", "above_class_average"]
    )
    ws1.append(ss_headers)
    for _, row in df.iterrows():
        sid = str(row["student_id"])
        ps  = per_student[sid]
        data_row = (
            [sid]
            + ps["quiz_scores"]
            + ps["test_scores"]
            + [
                ps["simple_average"],
                ps["weighted_average"],
                ps["letter_grade"],
                ps["slope"],
                ps["std_dev"],
                ps["highest_score"],
                ps["lowest_score"],
                ps["rank"],
                ps["percentile"],
                ps["z_score"],
                ps["above_class_average"],
            ]
        )
        ws1.append(data_row)

    # Sheet 2: Assessment Stats
    ws2 = wb.create_sheet("Assessment Stats")
    ws2.append(["assessment", "class_average", "std_dev", "highest_score", "lowest_score"])
    for col in all_score_cols:
        pa = per_assessment[col]
        ws2.append([col, pa["class_average"], pa["std_dev"], pa["highest_score"], pa["lowest_score"]])

    # Sheet 3: Z-Scores  (rows = students, cols = assessments)
    ws3 = wb.create_sheet("Z-Scores")
    ws3.append(["student_id"] + all_score_cols)
    for _, row in df.iterrows():
        sid = str(row["student_id"])
        z_row = [sid] + [per_assessment[col]["per_student_z_scores"][sid] for col in all_score_cols]
        ws3.append(z_row)

    # Sheet 4: Overall
    ws4 = wb.create_sheet("Overall")
    ws4.append(["class_weighted_average", overall["class_weighted_average"]])
    ws4.append(["class_median",           overall["class_median"]])
    for g in ALL_GRADES:
        ws4.append([f"grade_{g}", overall["grade_distribution"][g]])
    ws4.append(["most_improved_student",   overall["most_improved_student"]])
    ws4.append(["most_consistent_student", overall["most_consistent_student"]])

    OUTPUT_XLS.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLS)
    print(f"Written: {OUTPUT_XLS}")
    print(f"Sheets:  {wb.sheetnames}")

    # ── Sanity-check first student ────────────────────────────────────────────
    s = per_student["S001"]
    print(f"\nS001 sanity check:")
    print(f"  scores:           {s['raw_scores']}")
    print(f"  simple_average:   {s['simple_average']}")
    print(f"  weighted_average: {s['weighted_average']}  (expected 75.016667)")
    print(f"  slope:            {s['slope']}  (expected 3.781818)")
    print(f"  std_dev:          {s['std_dev']}  (expected 11.726513)")
    print(f"  letter_grade:     {s['letter_grade']}  (expected C)")
    print(f"  rank:             {s['rank']}  (expected 16)")
    print(f"  percentile:       {s['percentile']}  (expected 33.3)")
    print(f"  z_score:          {s['z_score']}  (expected -0.28)")
    print(f"  above_class_avg:  {s['above_class_average']}  (expected False)")

    # Tie check
    print(f"\nRank-tie check:")
    for sid, wa in sorted(student_was.items(), key=lambda x: -x[1]):
        print(f"  {sid}: wa={wa:.6f}  rank={ranks[sid]}")


if __name__ == "__main__":
    main()
