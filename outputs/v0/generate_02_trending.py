import csv
import math
import openpyxl

# ── Load CSV ──────────────────────────────────────────────────────────────────
INPUT_CSV  = "/Users/lucaschen/skillgraph/benchmark/test_inputs/02_trending.csv"
OUTPUT_XLSX = "/Users/lucaschen/skillgraph/outputs/v0/02_trending_output.xlsx"

students = []
with open(INPUT_CSV, newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        students.append({
            "student_id": row["student_id"],
            "scores": [float(row[f"test{i}"]) for i in range(1, 6)],
        })

# ── Helper functions ──────────────────────────────────────────────────────────
def ols_slope(y):
    """OLS slope with x = [1,2,3,4,5] per spec formula."""
    n = 5
    x = [1, 2, 3, 4, 5]
    sum_x  = sum(x)                         # 15
    sum_x2 = sum(xi**2 for xi in x)         # 55
    sum_y  = sum(y)
    sum_xy = sum(x[i]*y[i] for i in range(n))
    numerator   = n * sum_xy - sum_x * sum_y
    denominator = n * sum_x2 - sum_x**2     # always 50
    return numerator / denominator

def sample_std(y):
    """Sample standard deviation (ddof=1)."""
    n = len(y)
    mean = sum(y) / n
    return math.sqrt(sum((v - mean)**2 for v in y) / (n - 1))

def letter_grade(avg):
    if avg >= 96.5:  return "A+"
    if avg >= 92.5:  return "A"
    if avg >= 89.5:  return "A-"
    if avg >= 86.5:  return "B+"
    if avg >= 82.5:  return "B"
    if avg >= 79.5:  return "B-"
    if avg >= 76.5:  return "C+"
    if avg >= 72.5:  return "C"
    if avg >= 69.5:  return "C-"
    if avg >= 66.5:  return "D+"
    if avg >= 62.5:  return "D"
    if avg >= 59.5:  return "D-"
    return "F"

def r6(v):
    """Round to 6 decimal places."""
    return round(v, 6)

# ── Per-student stats ─────────────────────────────────────────────────────────
for s in students:
    y = s["scores"]
    s["average"]       = r6(sum(y) / 5)
    s["letter_grade"]  = letter_grade(s["average"])
    s["slope"]         = r6(ols_slope(y))
    s["std_dev"]       = r6(sample_std(y))
    s["highest_score"] = max(y)
    s["lowest_score"]  = min(y)

# ── Class average (mean of student averages) ──────────────────────────────────
class_avg = r6(sum(s["average"] for s in students) / len(students))

# ── Ranks: competition rank, descending by average ───────────────────────────
sorted_avgs = sorted((s["average"] for s in students), reverse=True)
def competition_rank(avg):
    return sorted_avgs.index(avg) + 1  # first occurrence = highest rank for tie

for s in students:
    s["rank"] = competition_rank(s["average"])
    s["above_class_average"] = bool(s["average"] > class_avg)

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Student Stats ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Student Stats"
headers1 = ["student_id","test1","test2","test3","test4","test5",
            "average","letter_grade","slope","std_dev",
            "highest_score","lowest_score","rank","above_class_average"]
ws1.append(headers1)
for s in students:
    ws1.append([
        s["student_id"],
        s["scores"][0], s["scores"][1], s["scores"][2], s["scores"][3], s["scores"][4],
        s["average"], s["letter_grade"], s["slope"], s["std_dev"],
        s["highest_score"], s["lowest_score"], s["rank"], s["above_class_average"],
    ])

# ── Sheet 2: Test Stats ───────────────────────────────────────────────────────
ws2 = wb.create_sheet("Test Stats")
headers2 = ["test","class_average","std_dev","highest_score","lowest_score"]
ws2.append(headers2)
for i in range(5):
    col_scores = [s["scores"][i] for s in students]
    n = len(col_scores)
    col_avg = sum(col_scores) / n
    ws2.append([
        f"test{i+1}",
        r6(col_avg),
        r6(sample_std(col_scores)),
        max(col_scores),
        min(col_scores),
    ])

# ── Sheet 3: Overall ──────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Overall")
ws3.append(["key", "value"])

all_avgs = [s["average"] for s in students]
n = len(all_avgs)
sorted_avgs_all = sorted(all_avgs)
if n % 2 == 1:
    class_median = sorted_avgs_all[n // 2]
else:
    class_median = (sorted_avgs_all[n//2 - 1] + sorted_avgs_all[n//2]) / 2

ws3.append(["class_average", r6(class_avg)])
ws3.append(["class_median",  r6(class_median)])

all_grades = ["A+","A","A-","B+","B","B-","C+","C","C-","D+","D","D-","F"]
grade_counts = {g: 0 for g in all_grades}
for s in students:
    grade_counts[s["letter_grade"]] += 1
for g in all_grades:
    ws3.append([f"grade_{g}", grade_counts[g]])

# most_improved: highest slope; tiebreak = higher average
most_improved = max(students, key=lambda s: (s["slope"], s["average"]))
ws3.append(["most_improved_student", most_improved["student_id"]])

# most_consistent: lowest std_dev; tiebreak = higher average
most_consistent = min(students, key=lambda s: (s["std_dev"], -s["average"]))
ws3.append(["most_consistent_student", most_consistent["student_id"]])

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_XLSX)
print(f"Wrote: {OUTPUT_XLSX}")
print(f"Sheets: {wb.sheetnames}")
