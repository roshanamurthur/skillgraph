#!/usr/bin/env python3
"""Generate 04_extreme_output.xlsx from 04_extreme.csv using the student grades skill spec."""

import csv
import math
from pathlib import Path

import openpyxl

# -- Paths ---------------------------------------------------------------------
INPUT_CSV   = "/Users/lucaschen/skillgraph/benchmark/test_inputs/04_extreme.csv"
OUTPUT_XLSX = "/Users/lucaschen/skillgraph/outputs/v0/04_extreme_output.xlsx"

# -- Weights -------------------------------------------------------------------
QUIZ_WEIGHT = 0.05          # 5% each, 4 quizzes = 20%
TEST_WEIGHT = 80 / 600      # 80% / 6 tests ≈ 13.3333% each

ASSESSMENT_NAMES = [
    "quiz_1", "quiz_2", "quiz_3", "quiz_4",
    "test_1", "test_2", "test_3", "test_4", "test_5", "test_6",
]

GRADE_SCALE = [
    (96.5, "A+"), (92.5, "A"),  (89.5, "A-"),
    (86.5, "B+"), (82.5, "B"),  (79.5, "B-"),
    (76.5, "C+"), (72.5, "C"),  (69.5, "C-"),
    (66.5, "D+"), (62.5, "D"),  (59.5, "D-"),
]

ALL_GRADES = ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]


# -- Helper functions ----------------------------------------------------------

def letter_grade(wa):
    for threshold, grade in GRADE_SCALE:
        if wa >= threshold:
            return grade
    return "F"


def ols_slope(y_values):
    """OLS slope with x = [1, 2, ..., n]."""
    n = len(y_values)
    x = list(range(1, n + 1))
    sum_x  = sum(x)
    sum_y  = sum(y_values)
    sum_xy = sum(xi * yi for xi, yi in zip(x, y_values))
    sum_x2 = sum(xi * xi for xi in x)
    denom  = n * sum_x2 - sum_x * sum_x
    return (n * sum_xy - sum_x * sum_y) / denom


def sample_std(values):
    """Sample standard deviation (ddof=1)."""
    n    = len(values)
    mean = sum(values) / n
    variance = sum((v - mean) ** 2 for v in values) / (n - 1)
    return math.sqrt(variance)


# -- Read CSV ------------------------------------------------------------------
students = []
with open(INPUT_CSV, newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        sid    = row["student_id"]
        scores = [float(row[a]) for a in ASSESSMENT_NAMES]
        students.append({"student_id": sid, "scores": scores})

n = len(students)

# -- Per-student base stats ----------------------------------------------------
for s in students:
    scores  = s["scores"]
    quizzes = scores[:4]
    tests   = scores[4:]

    s["simple_average"]   = round(sum(scores) / len(scores), 6)
    wa = QUIZ_WEIGHT * sum(quizzes) + TEST_WEIGHT * sum(tests)
    s["weighted_average"] = round(wa, 6)
    s["slope"]            = round(ols_slope(scores), 6)
    s["std_dev"]          = round(sample_std(scores), 6)
    s["highest_score"]    = max(scores)
    s["lowest_score"]     = min(scores)
    s["letter_grade"]     = letter_grade(s["weighted_average"])

# -- Class-level stats needed for rank / percentile / z_score -----------------
wa_values     = [s["weighted_average"] for s in students]
class_wa_mean = sum(wa_values) / n
class_wa_std  = sample_std(wa_values)

# -- Rank (competition rank, descending weighted_average) ---------------------
for s in students:
    wa        = s["weighted_average"]
    s["rank"] = sum(1 for other in wa_values if other > wa) + 1

# -- Percentile: (count strictly below / (n-1)) * 100 -------------------------
for s in students:
    wa              = s["weighted_average"]
    count_below     = sum(1 for other in wa_values if other < wa)
    s["percentile"] = round(count_below / (n - 1) * 100, 6)

# -- Student z_score: (wa - class_mean) / class_std, rounded 2dp --------------
for s in students:
    if class_wa_std == 0:
        s["z_score"] = 0.0
    else:
        z            = (s["weighted_average"] - class_wa_mean) / class_wa_std
        s["z_score"] = round(z, 2)

# -- above_class_average (strict >) -------------------------------------------
for s in students:
    s["above_class_average"] = s["weighted_average"] > class_wa_mean

# -- Assessment stats ----------------------------------------------------------
assessment_data = {}
for i, aname in enumerate(ASSESSMENT_NAMES):
    col_scores = [s["scores"][i] for s in students]
    col_mean   = sum(col_scores) / n
    col_std    = sample_std(col_scores)
    per_z      = {}
    for s in students:
        if col_std == 0:
            z = 0.0
        else:
            z = (s["scores"][i] - col_mean) / col_std
        per_z[s["student_id"]] = round(z, 2)
    assessment_data[aname] = {
        "class_average":        round(col_mean, 6),
        "std_dev":              round(col_std, 6),
        "highest_score":        max(col_scores),
        "lowest_score":         min(col_scores),
        "per_student_z_scores": per_z,
    }

# -- Overall stats -------------------------------------------------------------
class_wa_mean_r = round(class_wa_mean, 6)

sorted_wa = sorted(wa_values)
if n % 2 == 1:
    class_median = sorted_wa[n // 2]
else:
    class_median = (sorted_wa[n // 2 - 1] + sorted_wa[n // 2]) / 2
class_median = round(class_median, 6)

grade_dist = {g: 0 for g in ALL_GRADES}
for s in students:
    grade_dist[s["letter_grade"]] += 1

# Most improved: highest slope, tiebreak higher weighted_average
most_improved = max(students, key=lambda s: (s["slope"], s["weighted_average"]))
# Most consistent: lowest std_dev, tiebreak higher weighted_average
most_consistent = min(students, key=lambda s: (s["std_dev"], -s["weighted_average"]))


# -- EXECUTION TRACE LOG -------------------------------------------------------
print("\n=== EXECUTION TRACE ===\n")

print("WEIGHTED AVERAGE")
print(f"  Weights: quizzes=0.05 each (20% total), tests={TEST_WEIGHT:.10f} each (80% / 6 total)")
s0 = students[0]
print(f"  First student ({s0['student_id']}): scores={s0['scores']}")
print(f"    quiz sum={sum(s0['scores'][:4])}, test sum={sum(s0['scores'][4:])}")
print(f"    wa = 0.05*{sum(s0['scores'][:4])} + {TEST_WEIGHT:.10f}*{sum(s0['scores'][4:])} = {s0['weighted_average']}")

print("\nSLOPE FORMULA")
print("  Formula: OLS slope = (n·Σxy − Σx·Σy) / (n·Σx² − (Σx)²)")
print(f"  x-axis values: {list(range(1, 11))}  (x=1..10, one per assessment)")
y0 = s0["scores"]
x0 = list(range(1, 11))
n0 = 10
sx = sum(x0); sy = sum(y0)
sxy = sum(xi*yi for xi,yi in zip(x0,y0)); sx2 = sum(xi*xi for xi in x0)
slope_manual = (n0*sxy - sx*sy) / (n0*sx2 - sx*sx)
print(f"  {s0['student_id']}: Σx={sx}, Σy={sy}, Σxy={sxy}, Σx²={sx2}")
print(f"  slope = ({n0}*{sxy} - {sx}*{sy}) / ({n0}*{sx2} - {sx}²) = {slope_manual} → rounded={s0['slope']}")

print("\nSTANDARD DEVIATION")
print("  Using SAMPLE std dev (ddof=1)")
scores0 = s0["scores"]
mean0 = sum(scores0)/len(scores0)
var0 = sum((v-mean0)**2 for v in scores0) / (len(scores0)-1)
std0 = math.sqrt(var0)
print(f"  {s0['student_id']}: mean={mean0}, variance={var0}, std_dev={std0} → rounded={s0['std_dev']}")

print("\nLETTER GRADE BOUNDARIES (applied to weighted_average)")
print("  ≥96.5→A+ ≥92.5→A ≥89.5→A- ≥86.5→B+ ≥82.5→B ≥79.5→B- ≥76.5→C+ ≥72.5→C ≥69.5→C- ≥66.5→D+ ≥62.5→D ≥59.5→D- else→F")
BOUNDARY_PROXIMITY = 0.5
grade_boundaries = [96.5, 92.5, 89.5, 86.5, 82.5, 79.5, 76.5, 72.5, 69.5, 66.5, 62.5, 59.5]
for s in students:
    wa  = s["weighted_average"]
    near = [b for b in grade_boundaries if abs(wa - b) <= BOUNDARY_PROXIMITY]
    flag = f"  *** NEAR BOUNDARY {near} ***" if near else ""
    print(f"  {s['student_id']}: wa={wa:.6f} → {s['letter_grade']}{flag}")

print("\nRANK TIE-BREAKING")
print("  Strategy: competition rank (ties share highest rank; next rank skips)")
ties_map = {}
for s in students:
    r = s["rank"]
    ties_map.setdefault(r, []).append(s["student_id"])
tied_groups = {r: sids for r, sids in ties_map.items() if len(sids) > 1}
if tied_groups:
    print(f"  Tied groups: {tied_groups}")
else:
    print("  No ties in this dataset — all ranks are unique")
for s in students:
    print(f"  {s['student_id']}: wa={s['weighted_average']:.6f} → rank {s['rank']}")

print("\nPERCENTILE AND Z-SCORE")
print(f"  class_wa_mean = {class_wa_mean_r}")
print(f"  class_wa_sample_std (ddof=1) = {round(class_wa_std, 6)}")
print(f"  n = {n}")
print("  Percentile formula: (count strictly below / (n-1)) * 100")
print("  z_score formula: (student_wa - class_wa_mean) / class_wa_sample_std, rounded to 2dp")
print(f"  {s0['student_id']}: percentile={s0['percentile']}, z_score={s0['z_score']}")

print("\nINTERPRETATION CHOICES / EDGE CASES")
print("  - S001 (all 100s): perfect score → slope=0, std_dev=0, wa=100")
print("  - S002 (all 0s): all-zero scores → slope=0, std_dev=0, wa=0")
print("  - S003 (quiz=0, tests=100): wa = 0*4*0.05 + (80/600)*600 = 80 (tests only)")
print("  - S004 (quiz=100, tests=0): wa = 0.05*400 + 0 = 20 (quizzes only)")
print("  - Assessment std_dev: if col_std=0 (all same score), z_score set to 0.0 to avoid division by zero")
print("  - class_wa_std: if 0, student z_scores set to 0.0")
print("  - Percentile rounded to 6dp (matches other reference implementations)")

# Check for zero-std assessments
for aname in ASSESSMENT_NAMES:
    if assessment_data[aname]["std_dev"] == 0:
        print(f"  - Assessment {aname} has std_dev=0 (all scores identical) → per-student z=0.0")

print("\nOVERALL")
print(f"  class_weighted_average = {class_wa_mean_r}")
print(f"  class_median = {class_median}")
print(f"  grade_distribution = {grade_dist}")
print(f"  most_improved_student = {most_improved['student_id']} (slope={most_improved['slope']}, wa={most_improved['weighted_average']})")
print(f"  most_consistent_student = {most_consistent['student_id']} (std_dev={most_consistent['std_dev']}, wa={most_consistent['weighted_average']})")

print("\n=== END EXECUTION TRACE ===\n")


# -- Write Excel ---------------------------------------------------------------
Path(OUTPUT_XLSX).parent.mkdir(parents=True, exist_ok=True)
wb = openpyxl.Workbook()

# Sheet 1: Student Stats
ws1 = wb.active
ws1.title = "Student Stats"
headers1 = (
    ["student_id"] + ASSESSMENT_NAMES +
    ["simple_average", "weighted_average", "letter_grade", "slope", "std_dev",
     "highest_score", "lowest_score", "rank", "percentile", "z_score", "above_class_average"]
)
ws1.append(headers1)
for s in students:
    row = (
        [s["student_id"]] + s["scores"] +
        [s["simple_average"], s["weighted_average"], s["letter_grade"],
         s["slope"], s["std_dev"], s["highest_score"], s["lowest_score"],
         s["rank"], s["percentile"], s["z_score"], s["above_class_average"]]
    )
    ws1.append(row)

# Sheet 2: Assessment Stats
ws2 = wb.create_sheet("Assessment Stats")
ws2.append(["assessment", "class_average", "std_dev", "highest_score", "lowest_score"])
for aname in ASSESSMENT_NAMES:
    a = assessment_data[aname]
    ws2.append([aname, a["class_average"], a["std_dev"], a["highest_score"], a["lowest_score"]])

# Sheet 3: Z-Scores (rows=students, cols=assessments)
ws3 = wb.create_sheet("Z-Scores")
ws3.append(["student_id"] + ASSESSMENT_NAMES)
for s in students:
    row = [s["student_id"]] + [assessment_data[a]["per_student_z_scores"][s["student_id"]]
                                for a in ASSESSMENT_NAMES]
    ws3.append(row)

# Sheet 4: Overall
ws4 = wb.create_sheet("Overall")
ws4.append(["key", "value"])
ws4.append(["class_weighted_average", class_wa_mean_r])
ws4.append(["class_median",           class_median])
for g in ALL_GRADES:
    ws4.append([f"grade_{g}", grade_dist[g]])
ws4.append(["most_improved_student",   most_improved["student_id"]])
ws4.append(["most_consistent_student", most_consistent["student_id"]])

wb.save(OUTPUT_XLSX)
print(f"Output written: {OUTPUT_XLSX}")
print(f"Sheet names: {wb.sheetnames}")
