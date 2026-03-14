import csv
import math
import openpyxl

# ── Load input ──────────────────────────────────────────────────────────────
INPUT  = "/Users/lucaschen/skillgraph/benchmark/test_inputs/01_clean.csv"
OUTPUT = "/Users/lucaschen/skillgraph/outputs/v0/01_clean_output.xlsx"

students = []
with open(INPUT, newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        students.append({
            "student_id": row["student_id"],
            "scores": [float(row[f"test{i}"]) for i in range(1, 6)],
        })

# ── Helper functions ─────────────────────────────────────────────────────────

def ols_slope(y):
    """OLS slope with x = [1,2,3,4,5]."""
    n = 5
    x = [1, 2, 3, 4, 5]
    sum_x  = sum(x)
    sum_y  = sum(y)
    sum_xy = sum(xi * yi for xi, yi in zip(x, y))
    sum_x2 = sum(xi * xi for xi in x)
    return (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2)

def sample_std(values):
    """Sample standard deviation (ddof=1)."""
    n = len(values)
    mean = sum(values) / n
    return math.sqrt(sum((v - mean) ** 2 for v in values) / (n - 1))

def letter_grade(avg):
    if avg >= 96.5:  return "A+"
    if avg >= 92.5:  return "A"
    if avg >= 89.5:  return "A-"
    if avg >= 86.5:  return "B+"
    if avg >= 82.5:  return "B"
    if avg >= 79.5:  return "B-"
    if avg >= 76.5:  return "C+"
    if avg >= 72.5:  return "C"
    if avg >= 69.5:  return "C-"
    if avg >= 66.5:  return "D+"
    if avg >= 62.5:  return "D"
    if avg >= 59.5:  return "D-"
    return "F"

def r6(v):
    """Round to 6 decimal places."""
    return round(v, 6)

# ── Per-student stats ────────────────────────────────────────────────────────
for s in students:
    sc = s["scores"]
    s["average"]       = r6(sum(sc) / 5)
    s["letter_grade"]  = letter_grade(s["average"])
    s["slope"]         = r6(ols_slope(sc))
    s["std_dev"]       = r6(sample_std(sc))
    s["highest_score"] = max(sc)
    s["lowest_score"]  = min(sc)

# Class average
class_avg = r6(sum(s["average"] for s in students) / len(students))

for s in students:
    s["above_class_average"] = s["average"] > class_avg

# ── Rank (competition / "1224" style) ────────────────────────────────────────
sorted_avgs = sorted((s["average"] for s in students), reverse=True)
for s in students:
    s["rank"] = sorted_avgs.index(s["average"]) + 1

# ── Build Excel workbook ─────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Student Stats ───────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Student Stats"

headers1 = ["student_id","test1","test2","test3","test4","test5",
            "average","letter_grade","slope","std_dev",
            "highest_score","lowest_score","rank","above_class_average"]
ws1.append(headers1)

for s in students:
    ws1.append([
        s["student_id"],
        *[int(v) for v in s["scores"]],
        s["average"],
        s["letter_grade"],
        s["slope"],
        s["std_dev"],
        int(s["highest_score"]),
        int(s["lowest_score"]),
        s["rank"],
        s["above_class_average"],   # Python bool
    ])

# ── Sheet 2: Test Stats ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("Test Stats")

headers2 = ["test","class_average","std_dev","highest_score","lowest_score"]
ws2.append(headers2)

for i in range(1, 6):
    col_scores = [s["scores"][i-1] for s in students]
    ws2.append([
        f"test{i}",
        r6(sum(col_scores) / len(col_scores)),
        r6(sample_std(col_scores)),
        int(max(col_scores)),
        int(min(col_scores)),
    ])

# ── Sheet 3: Overall ─────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Overall")
ws3.append(["key", "value"])

averages = [s["average"] for s in students]
n = len(averages)
sorted_avgs_list = sorted(averages)
if n % 2 == 1:
    class_median = sorted_avgs_list[n // 2]
else:
    class_median = r6((sorted_avgs_list[n//2 - 1] + sorted_avgs_list[n//2]) / 2)

grade_buckets = ["A+","A","A-","B+","B","B-","C+","C","C-","D+","D","D-","F"]
grade_counts  = {g: sum(1 for s in students if s["letter_grade"] == g)
                 for g in grade_buckets}

# Most improved: highest slope; tiebreak = higher average
most_improved = max(students, key=lambda s: (s["slope"], s["average"]))

# Most consistent: lowest std_dev; tiebreak = higher average
most_consistent = min(students, key=lambda s: (s["std_dev"], -s["average"]))

ws3.append(["class_average", class_avg])
ws3.append(["class_median",  class_median])
for g in grade_buckets:
    ws3.append([f"grade_{g}", grade_counts[g]])
ws3.append(["most_improved_student",   most_improved["student_id"]])
ws3.append(["most_consistent_student", most_consistent["student_id"]])

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
