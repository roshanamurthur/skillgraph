"""Generate 04_extreme_output.xlsx per skill_v0.md specification."""
import math
import openpyxl

# ---------- raw data ----------
students = [
    {"student_id": "S001", "scores": [100, 100, 100, 100, 100]},
    {"student_id": "S002", "scores": [0, 0, 0, 0, 0]},
    {"student_id": "S003", "scores": [100, 0, 100, 0, 100]},
    {"student_id": "S004", "scores": [0, 100, 0, 100, 0]},
    {"student_id": "S005", "scores": [50, 50, 50, 50, 50]},
]

# ---------- helpers ----------
def ols_slope(scores):
    """OLS slope: (n·Σxy − Σx·Σy) / (n·Σx² − (Σx)²), x=[1..5]."""
    n = 5
    x = [1, 2, 3, 4, 5]
    sum_x  = sum(x)                         # 15
    sum_x2 = sum(xi**2 for xi in x)        # 55
    sum_y  = sum(scores)
    sum_xy = sum(x[i]*scores[i] for i in range(n))
    denom  = n * sum_x2 - sum_x**2         # 50
    return (n * sum_xy - sum_x * sum_y) / denom

def sample_std(scores):
    n = len(scores)
    mean = sum(scores) / n
    return math.sqrt(sum((s - mean)**2 for s in scores) / (n - 1))

def letter_grade(avg):
    if avg >= 96.5:  return "A+"
    if avg >= 92.5:  return "A"
    if avg >= 89.5:  return "A-"
    if avg >= 86.5:  return "B+"
    if avg >= 82.5:  return "B"
    if avg >= 79.5:  return "B-"
    if avg >= 76.5:  return "C+"
    if avg >= 72.5:  return "C"
    if avg >= 69.5:  return "C-"
    if avg >= 66.5:  return "D+"
    if avg >= 62.5:  return "D"
    if avg >= 59.5:  return "D-"
    return "F"

def r6(v):
    return round(v, 6)

# ---------- compute per-student stats ----------
for s in students:
    sc = s["scores"]
    s["average"]      = r6(sum(sc) / 5)
    s["letter_grade"] = letter_grade(s["average"])
    s["slope"]        = r6(ols_slope(sc))
    s["std_dev"]      = r6(sample_std(sc))
    s["highest_score"]= max(sc)
    s["lowest_score"] = min(sc)

# class average (mean of student averages)
class_avg = r6(sum(s["average"] for s in students) / len(students))

# above_class_average
for s in students:
    s["above_class_average"] = s["average"] > class_avg

# ranks: competition rank (ties share highest; next skips)
sorted_avgs = sorted((s["average"] for s in students), reverse=True)
def competition_rank(avg):
    return sorted_avgs.index(avg) + 1   # index of first occurrence = highest rank

for s in students:
    s["rank"] = competition_rank(s["average"])

# ---------- Sheet 1: Student Stats ----------
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Student Stats"

headers1 = ["student_id","test1","test2","test3","test4","test5",
            "average","letter_grade","slope","std_dev",
            "highest_score","lowest_score","rank","above_class_average"]
ws1.append(headers1)

for s in students:
    ws1.append([
        s["student_id"],
        s["scores"][0], s["scores"][1], s["scores"][2], s["scores"][3], s["scores"][4],
        s["average"], s["letter_grade"], s["slope"], s["std_dev"],
        s["highest_score"], s["lowest_score"], s["rank"], s["above_class_average"],
    ])

# ---------- Sheet 2: Test Stats ----------
ws2 = wb.create_sheet("Test Stats")
headers2 = ["test","class_average","std_dev","highest_score","lowest_score"]
ws2.append(headers2)

for i, test_name in enumerate(["test1","test2","test3","test4","test5"]):
    col = [s["scores"][i] for s in students]
    n = len(col)
    t_avg = sum(col) / n
    t_std = math.sqrt(sum((v - t_avg)**2 for v in col) / (n - 1))
    ws2.append([test_name, r6(t_avg), r6(t_std), max(col), min(col)])

# ---------- Sheet 3: Overall ----------
ws3 = wb.create_sheet("Overall")

avgs = [s["average"] for s in students]
sorted_avgs_med = sorted(avgs)
n = len(sorted_avgs_med)
class_median = r6((sorted_avgs_med[n//2 - 1] + sorted_avgs_med[n//2]) / 2
                  if n % 2 == 0 else sorted_avgs_med[n//2])

grade_buckets = ["A+","A","A-","B+","B","B-","C+","C","C-","D+","D","D-","F"]
grade_counts = {g: sum(1 for s in students if s["letter_grade"] == g) for g in grade_buckets}

# most_improved: highest slope; tiebreak = highest average
most_improved = max(students, key=lambda s: (s["slope"], s["average"]))["student_id"]

# most_consistent: lowest std_dev; tiebreak = highest average
most_consistent = min(students, key=lambda s: (s["std_dev"], -s["average"]))["student_id"]

ws3.append(["class_average",  r6(class_avg)])
ws3.append(["class_median",   class_median])
for g in grade_buckets:
    ws3.append([f"grade_{g}", grade_counts[g]])
ws3.append(["most_improved_student",   most_improved])
ws3.append(["most_consistent_student", most_consistent])

# ---------- save ----------
out_path = "/Users/lucaschen/skillgraph/outputs/v0/04_extreme_output.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheets: {wb.sheetnames}")
