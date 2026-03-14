import csv
import math
import openpyxl

INPUT_CSV  = "/Users/lucaschen/skillgraph/benchmark/test_inputs/03_ties.csv"
OUTPUT_XLSX = "/Users/lucaschen/skillgraph/outputs/v0/03_ties_output.xlsx"

QUIZ_COLS = ["quiz_1", "quiz_2", "quiz_3", "quiz_4"]
TEST_COLS = ["test_1", "test_2", "test_3", "test_4", "test_5", "test_6"]
ALL_COLS  = QUIZ_COLS + TEST_COLS  # x = [1..10]

QUIZ_WEIGHT = 0.05           # 5% each
TEST_WEIGHT = 80 / 600       # 80/6 % each

def ols_slope(scores):
    """OLS slope over x=[1..10], y=scores."""
    n = len(scores)
    x = list(range(1, n + 1))
    sx  = sum(x)
    sx2 = sum(xi**2 for xi in x)
    sy  = sum(scores)
    sxy = sum(x[i] * scores[i] for i in range(n))
    denom = n * sx2 - sx**2
    return (n * sxy - sx * sy) / denom

def sample_std(values):
    n = len(values)
    if n < 2:
        return 0.0
    mean = sum(values) / n
    var  = sum((v - mean)**2 for v in values) / (n - 1)
    return math.sqrt(var)

def letter_grade(wa):
    if wa >= 96.5: return "A+"
    if wa >= 92.5: return "A"
    if wa >= 89.5: return "A-"
    if wa >= 86.5: return "B+"
    if wa >= 82.5: return "B"
    if wa >= 79.5: return "B-"
    if wa >= 76.5: return "C+"
    if wa >= 72.5: return "C"
    if wa >= 69.5: return "C-"
    if wa >= 66.5: return "D+"
    if wa >= 62.5: return "D"
    if wa >= 59.5: return "D-"
    return "F"

def r6(x):
    return round(x, 6)

# ── Read CSV ──────────────────────────────────────────────────────────────────
students = []
with open(INPUT_CSV) as f:
    reader = csv.DictReader(f)
    for row in reader:
        sid    = row["student_id"]
        scores = [float(row[c]) for c in ALL_COLS]
        quizzes = scores[:4]
        tests   = scores[4:]
        wa = sum(q * QUIZ_WEIGHT for q in quizzes) + sum(t * TEST_WEIGHT for t in tests)
        sa = sum(scores) / len(scores)
        students.append({
            "student_id":       sid,
            "scores":           scores,
            "simple_average":   r6(sa),
            "weighted_average": r6(wa),
            "slope":            r6(ols_slope(scores)),
            "std_dev":          r6(sample_std(scores)),
            "highest_score":    max(scores),
            "lowest_score":     min(scores),
        })

# ── Letter grades ─────────────────────────────────────────────────────────────
for s in students:
    s["letter_grade"] = letter_grade(s["weighted_average"])

# ── Class-level weighted average stats ───────────────────────────────────────
was = [s["weighted_average"] for s in students]
n_students = len(students)
class_wa_mean = sum(was) / n_students
class_wa_std  = sample_std(was)  # sample std dev (ddof=1)

# ── Rank (dense rank by weighted_average descending) ─────────────────────────
# Ties share the same rank; next distinct value gets rank+1 (dense rank)
sorted_unique = sorted(set(was), reverse=True)
rank_map = {v: i + 1 for i, v in enumerate(sorted_unique)}
for s in students:
    s["rank"] = rank_map[s["weighted_average"]]

# ── Percentile: (count strictly below / (n-1)) * 100 ─────────────────────────
for s in students:
    count_below = sum(1 for wa in was if wa < s["weighted_average"])
    s["percentile"] = r6(count_below / (n_students - 1) * 100)

# ── Z-score: (wa - class_wa_mean) / class_wa_std ────────────────────────────
for s in students:
    s["z_score"] = r6((s["weighted_average"] - class_wa_mean) / class_wa_std)

# ── above_class_average (strict >) ───────────────────────────────────────────
for s in students:
    s["above_class_average"] = s["weighted_average"] > class_wa_mean

# ── Assessment stats ──────────────────────────────────────────────────────────
assessment_stats = []
for col in ALL_COLS:
    idx = ALL_COLS.index(col)
    col_scores = [s["scores"][idx] for s in students]
    col_mean   = sum(col_scores) / n_students
    col_std    = sample_std(col_scores)
    per_z = {}
    for s in students:
        sc = s["scores"][idx]
        per_z[s["student_id"]] = r6((sc - col_mean) / col_std) if col_std > 0 else 0.0
    assessment_stats.append({
        "assessment":    col,
        "class_average": r6(col_mean),
        "std_dev":       r6(col_std),
        "highest_score": max(col_scores),
        "lowest_score":  min(col_scores),
        "per_student_z": per_z,
    })

# ── Overall stats ─────────────────────────────────────────────────────────────
sorted_was = sorted(was)
if n_students % 2 == 0:
    class_median = (sorted_was[n_students//2 - 1] + sorted_was[n_students//2]) / 2
else:
    class_median = sorted_was[n_students // 2]

GRADES = ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]
grade_dist = {g: 0 for g in GRADES}
for s in students:
    grade_dist[s["letter_grade"]] += 1

# most_improved: highest slope, tiebreak highest weighted_average
max_slope = max(s["slope"] for s in students)
most_improved = max(
    (s for s in students if s["slope"] == max_slope),
    key=lambda s: s["weighted_average"]
)["student_id"]

# most_consistent: lowest std_dev, tiebreak highest weighted_average
min_std = min(s["std_dev"] for s in students)
most_consistent = max(
    (s for s in students if s["std_dev"] == min_std),
    key=lambda s: s["weighted_average"]
)["student_id"]

# ── Write Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Sheet 1: Student Stats
ws1 = wb.active
ws1.title = "Student Stats"
ws1.append([
    "student_id",
    *ALL_COLS,
    "simple_average", "weighted_average", "letter_grade",
    "slope", "std_dev", "highest_score", "lowest_score",
    "rank", "percentile", "z_score", "above_class_average",
])
for s in students:
    ws1.append([
        s["student_id"],
        *s["scores"],
        s["simple_average"],
        s["weighted_average"],
        s["letter_grade"],
        s["slope"],
        s["std_dev"],
        s["highest_score"],
        s["lowest_score"],
        s["rank"],
        s["percentile"],
        s["z_score"],
        s["above_class_average"],
    ])

# Sheet 2: Assessment Stats
ws2 = wb.create_sheet("Assessment Stats")
ws2.append(["assessment", "class_average", "std_dev", "highest_score", "lowest_score"])
for a in assessment_stats:
    ws2.append([
        a["assessment"],
        a["class_average"],
        a["std_dev"],
        a["highest_score"],
        a["lowest_score"],
    ])

# Sheet 3: Z-Scores (per student per assessment)
ws3 = wb.create_sheet("Z-Scores")
ws3.append(["student_id"] + ALL_COLS)
for s in students:
    row = [s["student_id"]]
    for a in assessment_stats:
        row.append(a["per_student_z"][s["student_id"]])
    ws3.append(row)

# Sheet 4: Overall
ws4 = wb.create_sheet("Overall")
ws4.append(["key", "value"])
ws4.append(["class_weighted_average", r6(class_wa_mean)])
ws4.append(["class_median",           r6(class_median)])
for g in GRADES:
    ws4.append([f"grade_{g}", grade_dist[g]])
ws4.append(["most_improved_student",   most_improved])
ws4.append(["most_consistent_student", most_consistent])

wb.save(OUTPUT_XLSX)
print(f"Written: {OUTPUT_XLSX}")
print(f"Sheets:  {wb.sheetnames}")
