import csv
import math
import openpyxl

INPUT_CSV = "/Users/lucaschen/skillgraph/benchmark/test_inputs/03_ties.csv"
OUTPUT_XLSX = "/Users/lucaschen/skillgraph/outputs/v0/03_ties_output.xlsx"

def compute_slope(scores):
    """OLS slope: (n*sum(x*y) - sum(x)*sum(y)) / (n*sum(x^2) - sum(x)^2), x=[1..5]"""
    n = 5
    x = [1, 2, 3, 4, 5]
    sum_x = sum(x)          # 15
    sum_x2 = sum(xi**2 for xi in x)  # 55
    sum_y = sum(scores)
    sum_xy = sum(x[i] * scores[i] for i in range(n))
    denom = n * sum_x2 - sum_x**2  # 50
    return (n * sum_xy - sum_x * sum_y) / denom

def compute_sample_std(scores):
    """Sample standard deviation (ddof=1)."""
    n = len(scores)
    mean = sum(scores) / n
    variance = sum((s - mean)**2 for s in scores) / (n - 1)
    return math.sqrt(variance)

def letter_grade(avg):
    if avg >= 96.5: return "A+"
    elif avg >= 92.5: return "A"
    elif avg >= 89.5: return "A-"
    elif avg >= 86.5: return "B+"
    elif avg >= 82.5: return "B"
    elif avg >= 79.5: return "B-"
    elif avg >= 76.5: return "C+"
    elif avg >= 72.5: return "C"
    elif avg >= 69.5: return "C-"
    elif avg >= 66.5: return "D+"
    elif avg >= 62.5: return "D"
    elif avg >= 59.5: return "D-"
    else: return "F"

def competition_rank(values):
    """Returns dict: value -> competition rank (ties share highest rank, next skips)."""
    sorted_vals = sorted(set(values), reverse=True)
    rank_map = {}
    current_rank = 1
    for v in sorted_vals:
        count = values.count(v)
        rank_map[v] = current_rank
        current_rank += count
    return rank_map

def r6(x):
    return round(x, 6)

# Read CSV
students = []
with open(INPUT_CSV) as f:
    reader = csv.DictReader(f)
    for row in reader:
        sid = row["student_id"]
        scores = [float(row[f"test{i}"]) for i in range(1, 6)]
        students.append({"student_id": sid, "scores": scores})

# Compute per-student stats
for s in students:
    scores = s["scores"]
    s["average"] = r6(sum(scores) / 5)
    s["letter_grade"] = letter_grade(s["average"])
    s["slope"] = r6(compute_slope(scores))
    s["std_dev"] = r6(compute_sample_std(scores))
    s["highest_score"] = max(scores)
    s["lowest_score"] = min(scores)

# Class average
class_avg = r6(sum(s["average"] for s in students) / len(students))

# Ranks (competition rank by average descending)
averages = [s["average"] for s in students]
rank_map = competition_rank(averages)
for s in students:
    s["rank"] = rank_map[s["average"]]
    s["above_class_average"] = s["average"] > class_avg

# Write Excel
wb = openpyxl.Workbook()

# ── Sheet 1: Student Stats ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Student Stats"
headers1 = ["student_id", "test1", "test2", "test3", "test4", "test5",
            "average", "letter_grade", "slope", "std_dev",
            "highest_score", "lowest_score", "rank", "above_class_average"]
ws1.append(headers1)
for s in students:
    ws1.append([
        s["student_id"],
        s["scores"][0], s["scores"][1], s["scores"][2], s["scores"][3], s["scores"][4],
        s["average"],
        s["letter_grade"],
        s["slope"],
        s["std_dev"],
        s["highest_score"],
        s["lowest_score"],
        s["rank"],
        s["above_class_average"],
    ])

# ── Sheet 2: Test Stats ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Test Stats")
headers2 = ["test", "class_average", "std_dev", "highest_score", "lowest_score"]
ws2.append(headers2)
for i in range(1, 6):
    test_scores = [s["scores"][i-1] for s in students]
    n = len(test_scores)
    mean = sum(test_scores) / n
    std = compute_sample_std(test_scores)
    ws2.append([
        f"test{i}",
        r6(mean),
        r6(std),
        max(test_scores),
        min(test_scores),
    ])

# ── Sheet 3: Overall ────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Overall")
ws3.append(["key", "value"])

avg_list = sorted([s["average"] for s in students])
n_s = len(avg_list)
class_median = (avg_list[n_s//2 - 1] + avg_list[n_s//2]) / 2 if n_s % 2 == 0 else avg_list[n_s//2]

grades_order = ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]
grade_counts = {g: 0 for g in grades_order}
for s in students:
    grade_counts[s["letter_grade"]] += 1

# most_improved: highest slope, tiebreak by highest average
max_slope = max(s["slope"] for s in students)
candidates_improved = [s for s in students if s["slope"] == max_slope]
most_improved = max(candidates_improved, key=lambda s: s["average"])["student_id"]

# most_consistent: lowest std_dev, tiebreak by highest average
min_std = min(s["std_dev"] for s in students)
candidates_consistent = [s for s in students if s["std_dev"] == min_std]
most_consistent = max(candidates_consistent, key=lambda s: s["average"])["student_id"]

ws3.append(["class_average", r6(class_avg)])
ws3.append(["class_median", r6(class_median)])
for g in grades_order:
    ws3.append([f"grade_{g}", grade_counts[g]])
ws3.append(["most_improved_student", most_improved])
ws3.append(["most_consistent_student", most_consistent])

wb.save(OUTPUT_XLSX)
print(f"Written: {OUTPUT_XLSX}")
print(f"Sheets: {wb.sheetnames}")
