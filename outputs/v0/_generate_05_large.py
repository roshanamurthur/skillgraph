#!/usr/bin/env python3
"""Generate 05_large_output.xlsx from 05_large.csv."""

import csv
import math
import openpyxl

INPUT_CSV  = "/Users/lucaschen/skillgraph/benchmark/test_inputs/05_large.csv"
OUTPUT_XLSX = "/Users/lucaschen/skillgraph/outputs/v0/05_large_output.xlsx"

QUIZ_COLS = ["quiz_1", "quiz_2", "quiz_3", "quiz_4"]
TEST_COLS = ["test_1", "test_2", "test_3", "test_4", "test_5", "test_6"]
ALL_COLS  = QUIZ_COLS + TEST_COLS

QUIZ_WEIGHT = 0.05           # 5% each × 4 = 20%
TEST_WEIGHT = 0.80 / 6       # 80%/6 each × 6 = 80%

GRADE_SCALE = [
    (96.5, "A+"), (92.5, "A"), (89.5, "A-"), (86.5, "B+"),
    (82.5, "B"),  (79.5, "B-"), (76.5, "C+"), (72.5, "C"),
    (69.5, "C-"), (66.5, "D+"), (62.5, "D"),  (59.5, "D-"),
]

def read_csv(path):
    rows = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            entry = {"student_id": row["student_id"]}
            for col in ALL_COLS:
                entry[col] = float(row[col])
            rows.append(entry)
    return rows

def weighted_avg(s):
    q = sum(s[c] for c in QUIZ_COLS)
    t = sum(s[c] for c in TEST_COLS)
    return round(QUIZ_WEIGHT * q + TEST_WEIGHT * t, 6)

def simple_avg(s):
    vals = [s[c] for c in ALL_COLS]
    return round(sum(vals) / len(vals), 6)

def ols_slope(s):
    y = [s[c] for c in ALL_COLS]
    n = len(y)
    x = list(range(1, n + 1))
    sx  = sum(x)
    sy  = sum(y)
    sxy = sum(xi * yi for xi, yi in zip(x, y))
    sx2 = sum(xi ** 2 for xi in x)
    return round((n * sxy - sx * sy) / (n * sx2 - sx ** 2), 6)

def sample_std(values):
    n = len(values)
    if n <= 1:
        return 0.0
    mean = sum(values) / n
    var  = sum((v - mean) ** 2 for v in values) / (n - 1)
    return round(math.sqrt(var), 6)

def letter_grade(wa):
    for threshold, grade in GRADE_SCALE:
        if wa >= threshold:
            return grade
    return "F"

def percentile_formula(wa, all_was, n):
    count_below = sum(1 for w in all_was if w < wa)
    return round((count_below / (n - 1)) * 100, 1)

def main():
    students = read_csv(INPUT_CSV)
    n = len(students)

    # -------------------------------------------------------------------------
    # Per-student computations
    # -------------------------------------------------------------------------
    for s in students:
        s["simple_average"]   = simple_avg(s)
        s["weighted_average"] = weighted_avg(s)
        s["slope"]            = ols_slope(s)
        vals = [s[c] for c in ALL_COLS]
        s["std_dev"]          = sample_std(vals)
        s["highest_score"]    = max(vals)
        s["lowest_score"]     = min(vals)
        s["letter_grade"]     = letter_grade(s["weighted_average"])

    all_was = [s["weighted_average"] for s in students]
    sorted_was_desc = sorted(all_was, reverse=True)

    # Rank (competition rank, ties share highest)
    for s in students:
        s["rank"] = sorted_was_desc.index(s["weighted_average"]) + 1

    # Class weighted average stats (for z-score and above_class_average)
    class_wa_mean = sum(all_was) / n
    class_wa_std  = sample_std(all_was)

    for s in students:
        s["percentile"] = percentile_formula(s["weighted_average"], all_was, n)
        s["z_score"]    = round((s["weighted_average"] - class_wa_mean) / class_wa_std, 2)
        s["above_class_average"] = s["weighted_average"] > class_wa_mean

    # -------------------------------------------------------------------------
    # DECISION LOG: LETTER GRADE BOUNDARIES
    # -------------------------------------------------------------------------
    BOUNDARY_FLAGS = [96.5, 92.5, 89.5, 86.5, 82.5, 79.5, 76.5, 72.5, 69.5, 66.5, 62.5, 59.5]
    print("LETTER GRADE BOUNDARIES")
    print(f"  Thresholds: A+(96.5) A(92.5) A-(89.5) B+(86.5) B(82.5) B-(79.5) "
          f"C+(76.5) C(72.5) C-(69.5) D+(66.5) D(62.5) D-(59.5) F")
    for s in students:
        wa = s["weighted_average"]
        grade = s["letter_grade"]
        near = any(abs(wa - b) <= 0.5 for b in BOUNDARY_FLAGS)
        flag = " *** BOUNDARY PROXIMITY ***" if near else ""
        print(f"  {s['student_id']}: WA={wa:.6f} → {grade}{flag}")

    # -------------------------------------------------------------------------
    # DECISION LOG: RANK TIE-BREAKING
    # -------------------------------------------------------------------------
    print("\nRANK TIE-BREAKING")
    print("  Strategy: Competition rank (tied students share highest rank; next rank skips)")
    wa_counts = {}
    for wa in all_was:
        wa_counts[wa] = wa_counts.get(wa, 0) + 1
    tied_found = False
    for s in students:
        if wa_counts[s["weighted_average"]] > 1:
            print(f"  TIE: {s['student_id']} WA={s['weighted_average']:.6f} → rank {s['rank']}")
            tied_found = True
    if not tied_found:
        print("  No ties exist in this dataset.")

    # -------------------------------------------------------------------------
    # DECISION LOG: STANDARD DEVIATION (first student)
    # -------------------------------------------------------------------------
    s0 = students[0]
    print(f"\nSTANDARD DEVIATION")
    print(f"  Using sample std dev (ddof=1) over all 10 scores")
    print(f"  {s0['student_id']}: std_dev = {s0['std_dev']:.6f}")

    # -------------------------------------------------------------------------
    # DECISION LOG: SLOPE (first student)
    # -------------------------------------------------------------------------
    print(f"\nSLOPE FORMULA")
    print(f"  OLS: slope = (n·Σ(x·y) − Σx·Σy) / (n·Σ(x²) − (Σx)²)")
    print(f"  x-axis: [1,2,3,4,5,6,7,8,9,10] (quiz_1..quiz_4, test_1..test_6)")
    print(f"  {s0['student_id']}: slope = {s0['slope']:.6f}")

    # -------------------------------------------------------------------------
    # DECISION LOG: PERCENTILE & Z-SCORE
    # -------------------------------------------------------------------------
    print(f"\nPERCENTILE AND Z-SCORE")
    print(f"  Percentile: (count strictly below / (n-1)) * 100, rounded to 1 decimal, n={n}")
    print(f"  Z-score: (student_wa - class_wa_mean) / class_wa_sample_std, rounded to 2 decimals")
    print(f"  Class WA mean = {class_wa_mean:.6f}, Class WA sample std = {class_wa_std:.6f}")
    print(f"  {s0['student_id']}: percentile={s0['percentile']}, z_score={s0['z_score']}")

    # -------------------------------------------------------------------------
    # Per-assessment computations
    # -------------------------------------------------------------------------
    assessments = {}
    for col in ALL_COLS:
        vals = [s[col] for s in students]
        col_mean = sum(vals) / n
        col_std  = sample_std(vals)
        per_student_z = {}
        for s in students:
            per_student_z[s["student_id"]] = round((s[col] - col_mean) / col_std, 2)
        assessments[col] = {
            "class_average":        round(col_mean, 6),
            "std_dev":              col_std,
            "highest_score":        max(vals),
            "lowest_score":         min(vals),
            "per_student_z_scores": per_student_z,
        }

    # -------------------------------------------------------------------------
    # Overall stats
    # -------------------------------------------------------------------------
    sorted_was = sorted(all_was)
    if n % 2 == 1:
        class_median = sorted_was[n // 2]
    else:
        class_median = (sorted_was[n // 2 - 1] + sorted_was[n // 2]) / 2

    grade_dist = {}
    for grade_name in ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]:
        grade_dist[grade_name] = sum(1 for s in students if s["letter_grade"] == grade_name)

    # Most improved: highest slope, tiebreaker = higher WA
    most_improved = max(students, key=lambda s: (s["slope"], s["weighted_average"]))
    # Most consistent: lowest std_dev, tiebreaker = higher WA
    most_consistent = min(students, key=lambda s: (s["std_dev"], -s["weighted_average"]))

    # -------------------------------------------------------------------------
    # Write Excel
    # -------------------------------------------------------------------------
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Student Stats ----
    ws1 = wb.active
    ws1.title = "Student Stats"
    headers1 = (
        ["student_id"] + ALL_COLS +
        ["simple_average", "weighted_average", "letter_grade", "slope",
         "std_dev", "highest_score", "lowest_score", "rank",
         "percentile", "z_score", "above_class_average"]
    )
    ws1.append(headers1)
    for s in students:
        row = ([s["student_id"]] + [s[c] for c in ALL_COLS] + [
            round(s["simple_average"],   6),
            round(s["weighted_average"], 6),
            s["letter_grade"],
            round(s["slope"],            6),
            round(s["std_dev"],          6),
            s["highest_score"],
            s["lowest_score"],
            s["rank"],
            s["percentile"],
            s["z_score"],
            s["above_class_average"],
        ])
        ws1.append(row)

    # ---- Sheet 2: Assessment Stats ----
    ws2 = wb.create_sheet("Assessment Stats")
    headers2 = ["assessment", "class_average", "std_dev", "highest_score", "lowest_score"]
    ws2.append(headers2)
    for col in ALL_COLS:
        a = assessments[col]
        ws2.append([
            col,
            round(a["class_average"], 6),
            round(a["std_dev"],       6),
            a["highest_score"],
            a["lowest_score"],
        ])

    # ---- Sheet 3: Z-Scores ----
    ws3 = wb.create_sheet("Z-Scores")
    headers3 = ["student_id"] + ALL_COLS
    ws3.append(headers3)
    for s in students:
        row = [s["student_id"]] + [assessments[col]["per_student_z_scores"][s["student_id"]] for col in ALL_COLS]
        ws3.append(row)

    # ---- Sheet 4: Overall ----
    ws4 = wb.create_sheet("Overall")
    overall_rows = [
        ("class_weighted_average",  round(class_wa_mean, 6)),
        ("class_median",            round(class_median,  6)),
    ]
    for grade_name in ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]:
        overall_rows.append((f"grade_{grade_name}", grade_dist[grade_name]))
    overall_rows.append(("most_improved_student",   most_improved["student_id"]))
    overall_rows.append(("most_consistent_student", most_consistent["student_id"]))
    for key, val in overall_rows:
        ws4.append([key, val])

    wb.save(OUTPUT_XLSX)
    print(f"\nOutput written: {OUTPUT_XLSX}")
    print(f"Sheets: {wb.sheetnames}")

    # Summary
    print(f"\nINTERPRETATION CHOICES")
    print(f"  - Skill v0.md describes test1-test5, but input has quiz_1-4 + test_1-6.")
    print(f"    Used weighted average formula from benchmark prompt (not v0 spec).")
    print(f"  - above_class_average uses strict > (not >=).")
    print(f"  - z_scores rounded to 2 decimal places; percentile to 1 decimal place.")
    print(f"  - other floats to 6 decimal places.")
    print(f"  - most_improved: highest slope, tiebreaker = higher WA.")
    print(f"  - most_consistent: lowest std_dev, tiebreaker = higher WA.")
    print(f"\nEDGE CASES")
    perfect = [s["student_id"] for s in students if any(s[c] == 100.0 for c in ALL_COLS)]
    print(f"  Students with at least one perfect score: {perfect}")
    print(f"\nMost improved: {most_improved['student_id']} (slope={most_improved['slope']:.6f})")
    print(f"Most consistent: {most_consistent['student_id']} (std_dev={most_consistent['std_dev']:.6f})")
    print(f"Class WA mean: {class_wa_mean:.6f}, Class WA sample std: {class_wa_std:.6f}")
    print(f"Class median WA: {round(class_median, 6):.6f}")
    print(f"Grade distribution: {grade_dist}")

if __name__ == "__main__":
    main()
