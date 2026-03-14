#!/usr/bin/env python3
"""
run_benchmark.py  --  Benchmark runner for the student grades skill.

Usage:
    python3 run_benchmark.py --version v0

For each test input CSV it:
  1. Invokes the skill via the Claude CLI
  2. Parses the generated Excel output
  3. Compares every computed value against the ground truth JSON
  4. Writes a scored JSON report with exact cell locations per failure
  5. Prints a human-readable summary table with failure locations
  6. Diffs against the previous version's results if available
"""

import argparse
import json
import re
import subprocess
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# -- Paths ---------------------------------------------------------------------
ROOT          = Path(__file__).parent.parent
BENCHMARK_DIR = Path(__file__).parent
INPUTS_DIR    = BENCHMARK_DIR / "test_inputs"
GT_DIR        = BENCHMARK_DIR / "ground_truth"
SKILL_DIR     = ROOT / "skill"
LOGS_BASE     = ROOT / "logs" / "eval_results"
LOGS_BASE.mkdir(parents=True, exist_ok=True)

TOLERANCE = 0.001

GRADE_BOUNDARIES = [96.5, 92.5, 89.5, 86.5, 82.5, 79.5, 76.5, 72.5, 69.5, 66.5, 62.5, 59.5]
BOUNDARY_PROXIMITY = 0.5


# -- Helpers -------------------------------------------------------------------

def _close_enough(actual, expected, tol=TOLERANCE):
    try:
        return abs(float(actual) - float(expected)) <= tol
    except (TypeError, ValueError):
        return str(actual).strip() == str(expected).strip()

def _delta(actual, expected):
    try:
        return round(abs(float(actual) - float(expected)), 6)
    except (TypeError, ValueError):
        return None

def fraction(passed, total):
    return f"{passed}/{total}"

def score(passed, total):
    return round(passed / total, 2) if total > 0 else 0.0

def norm(s):
    return re.sub(r"[\s_]+", "_", str(s).strip().lower())

def make_category(passed_checks, failed_checks):
    p = len(passed_checks)
    t = p + len(failed_checks)
    return {
        "fraction":      fraction(p, t),
        "score":         score(p, t),
        "passed":        p,
        "total":         t,
        "passed_checks": passed_checks,   # {check, actual} only
        "failed_checks": failed_checks,   # full location + delta
    }


# -- Failure descriptions ------------------------------------------------------

def describe_failure(check_name, actual, expected, delta=None):
    delta_str = f" — delta {delta}" if delta is not None else ""
    parts = check_name.split(".")

    if parts[0] == "student" and len(parts) == 3:
        sid, field = parts[1], parts[2]
        if field == "weighted_average":
            hint = "suggests simple average used instead of weighted" if delta and delta > 0.1 else "check weight formula"
            return f"{sid} weighted_average: got {actual}, expected {expected}{delta_str} — {hint}"
        if field == "simple_average":
            return f"{sid} simple_average: got {actual}, expected {expected}{delta_str} — check mean calculation"
        if field == "letter_grade":
            return f"{sid} letter_grade: got '{actual}', expected '{expected}' — likely wrong grade scale or based on simple avg instead of weighted avg"
        if field == "std_dev":
            return f"{sid} std_dev: got {actual}, expected {expected}{delta_str} — likely using population (ddof=0) instead of sample (ddof=1)"
        if field == "slope":
            return f"{sid} slope: got {actual}, expected {expected}{delta_str} — check OLS formula or score ordering (quiz_1..quiz_4, test_1..test_6)"
        if field == "rank":
            return f"{sid} rank: got {actual}, expected {expected} — check tie-handling (tied students share highest rank)"
        if field == "percentile":
            return f"{sid} percentile: got {actual}, expected {expected}{delta_str} — formula: (count strictly below / (n-1)) * 100"
        if field == "z_score":
            return f"{sid} z_score: got {actual}, expected {expected}{delta_str} — z = (student_wa - class_wa_mean) / class_wa_sample_std"
        if field == "above_class_average":
            return f"{sid} above_class_average: got {actual}, expected {expected} — check class average or use strict > not >="
        if field in ("highest_score", "lowest_score"):
            return f"{sid} {field}: got {actual}, expected {expected} — check min/max over all score columns"
        return f"{sid} {field}: got {actual}, expected {expected}{delta_str}"

    if parts[0] == "assessment" and len(parts) == 3:
        aname, field = parts[1], parts[2]
        if field == "std_dev":
            return f"{aname} std_dev: got {actual}, expected {expected}{delta_str} — use sample std dev (ddof=1)"
        if field == "class_average":
            return f"{aname} class_average: got {actual}, expected {expected}{delta_str} — check column mean"
        if field in ("highest_score", "lowest_score"):
            return f"{aname} {field}: got {actual}, expected {expected} — check min/max for this column"
        return f"{aname} {field}: got {actual}, expected {expected}{delta_str}"

    if parts[0] == "assessment_z" and len(parts) == 3:
        aname, sid = parts[1], parts[2]
        return f"{aname} z_score for {sid}: got {actual}, expected {expected}{delta_str} — z = (score - col_mean) / col_std"

    if parts[0] == "overall":
        field = ".".join(parts[1:])
        if field == "class_weighted_average":
            return f"class_weighted_average: got {actual}, expected {expected}{delta_str} — should be mean of student weighted_averages"
        if field == "class_median":
            return f"class_median: got {actual}, expected {expected}{delta_str} — should be median of student weighted_averages"
        if field == "most_improved_student":
            return f"most_improved_student: got '{actual}', expected '{expected}' — student with highest OLS slope"
        if field == "most_consistent_student":
            return f"most_consistent_student: got '{actual}', expected '{expected}' — student with lowest sample std dev"
        if field.startswith("grade_dist."):
            g = field.split(".")[-1]
            return f"grade_distribution['{g}']: got {actual}, expected {expected} — check grade boundary for '{g}'"
        return f"overall {field}: got {actual}, expected {expected}{delta_str}"

    return f"{check_name}: got {actual}, expected {expected}{delta_str}"


# -- Boundary-case detector ----------------------------------------------------

def is_boundary_check(check_name, gt):
    parts = check_name.split(".")
    if parts[0] != "student" or len(parts) != 3:
        return False
    sid, field = parts[1], parts[2]
    gt_s = gt["per_student"].get(sid, {})

    if field == "rank":
        expected_rank = gt_s.get("rank")
        if sum(1 for s in gt["per_student"].values() if s["rank"] == expected_rank) > 1:
            return True

    if field == "percentile":
        expected_pct = gt_s.get("percentile")
        if sum(1 for s in gt["per_student"].values() if s["percentile"] == expected_pct) > 1:
            return True

    if field in ("letter_grade", "weighted_average", "percentile", "z_score"):
        wa = gt_s.get("weighted_average", 0)
        if any(abs(wa - b) <= BOUNDARY_PROXIMITY for b in GRADE_BOUNDARIES):
            return True

    scores = gt_s.get("scores", [])
    if scores and (0.0 in scores or 100.0 in scores):
        return True

    if field == "std_dev" and scores and len(set(scores)) == 1:
        return True

    return False


# -- Location resolution -------------------------------------------------------

def resolve_location(check_name, location_map):
    """
    Map a check name to its exact Excel location.
    Returns a dict with: cell, sheet, row, column_letter, column_name, entity
    or None if the location cannot be resolved.
    """
    if location_map is None:
        return None
    parts = check_name.split(".")

    def lookup_row(lm, key):
        rows = lm.get("rows", {})
        return rows.get(key) or rows.get(norm(key))

    def lookup_col(lm, field):
        cols = lm.get("columns", {})
        return cols.get(field) or cols.get(norm(field))

    if parts[0] == "student" and len(parts) == 3:
        sid, field = parts[1], parts[2]
        sheet = "Student Stats"
        lm = location_map.get(sheet, {})
        row = lookup_row(lm, sid)
        col = lookup_col(lm, field)
        if row and col:
            return {"cell": f"'{sheet}'!{col}{row}", "sheet": sheet,
                    "row": row, "column_letter": col, "column_name": field, "entity": sid}

    if parts[0] == "assessment" and len(parts) == 3:
        aname, field = parts[1], parts[2]
        sheet = "Assessment Stats"
        lm = location_map.get(sheet, {})
        row = lookup_row(lm, aname)
        col = lookup_col(lm, field)
        if row and col:
            return {"cell": f"'{sheet}'!{col}{row}", "sheet": sheet,
                    "row": row, "column_letter": col, "column_name": field, "entity": aname}

    if parts[0] == "assessment_z" and len(parts) == 3:
        aname, sid = parts[1], parts[2]
        sheet = "Z-Scores"
        lm = location_map.get(sheet, {})
        row = lookup_row(lm, sid)
        col = lookup_col(lm, aname)
        if row and col:
            return {"cell": f"'{sheet}'!{col}{row}", "sheet": sheet,
                    "row": row, "column_letter": col,
                    "column_name": aname, "entity": f"{sid}/{aname}"}

    if parts[0] == "overall":
        field = ".".join(parts[1:])
        sheet = "Overall"
        lm = location_map.get(sheet, {})
        # grade_dist.A+ → look for row keyed as "grade_a+"
        if field.startswith("grade_dist."):
            grade = field.split(".")[-1]
            row_key = norm(f"grade_{grade}")
        else:
            row_key = field
        row = lookup_row(lm, row_key)
        col = lm.get("columns", {}).get("value", "B")
        if row:
            return {"cell": f"'{sheet}'!{col}{row}", "sheet": sheet,
                    "row": row, "column_letter": col, "column_name": field, "entity": "Overall"}

    return None


# -- Comparison engine ---------------------------------------------------------

def compare(parsed, gt, location_map=None):
    student_pass, student_fail = [], []
    assess_pass,  assess_fail  = [], []
    overall_pass, overall_fail = [], []
    advanced_pass, advanced_fail = [], []
    boundary_pass, boundary_fail = [], []

    def record(p_list, f_list, name, actual, expected, category, context=""):
        ok = _close_enough(actual, expected)
        if ok:
            p_list.append({"check": name, "actual": actual})
        else:
            d = _delta(actual, expected)
            entry = {
                "check":       name,
                "passed":      False,
                "actual":      actual,
                "expected":    expected,
                "delta":       d,
                "description": describe_failure(name, actual, expected, d),
            }
            loc = resolve_location(name, location_map)
            if loc:
                entry.update(loc)
            elif context:
                entry["entity"] = context
            f_list.append(entry)
        return ok

    base_fields = [
        "simple_average", "weighted_average", "letter_grade", "slope",
        "std_dev", "highest_score", "lowest_score", "rank", "above_class_average",
    ]
    advanced_student_fields = ["percentile", "z_score"]

    for sid, gt_s in gt["per_student"].items():
        p_s = parsed["per_student"].get(sid, {})
        for field in base_fields:
            name = f"student.{sid}.{field}"
            if is_boundary_check(name, gt):
                record(boundary_pass, boundary_fail, name, p_s.get(field), gt_s[field], "boundary_cases", sid)
            else:
                record(student_pass, student_fail, name, p_s.get(field), gt_s[field], "student_calculations", sid)
        for field in advanced_student_fields:
            name = f"student.{sid}.{field}"
            if is_boundary_check(name, gt):
                record(boundary_pass, boundary_fail, name, p_s.get(field), gt_s[field], "boundary_cases", sid)
            else:
                record(advanced_pass, advanced_fail, name, p_s.get(field), gt_s[field], "advanced_metrics", sid)

    for aname, gt_a in gt["per_assessment"].items():
        p_a = parsed["per_assessment"].get(aname, {})
        for field in ["class_average", "std_dev", "highest_score", "lowest_score"]:
            record(assess_pass, assess_fail, f"assessment.{aname}.{field}",
                   p_a.get(field), gt_a[field], "assessment_stats", aname)

    for aname, gt_a in gt["per_assessment"].items():
        gt_zs = gt_a.get("per_student_z_scores", {})
        p_zs  = parsed["per_assessment"].get(aname, {}).get("per_student_z_scores", {})
        for sid, expected_z in gt_zs.items():
            record(advanced_pass, advanced_fail, f"assessment_z.{aname}.{sid}",
                   p_zs.get(sid), expected_z, "advanced_metrics", f"{aname}/{sid}")

    gt_ov = gt["overall"]
    p_ov  = parsed["overall"]
    for field in ["class_weighted_average", "class_median",
                  "most_improved_student", "most_consistent_student"]:
        record(overall_pass, overall_fail, f"overall.{field}",
               p_ov.get(field), gt_ov[field], "overall_stats")
    for grade, gt_count in gt_ov["grade_distribution"].items():
        record(overall_pass, overall_fail, f"overall.grade_dist.{grade}",
               p_ov.get("grade_distribution", {}).get(grade), gt_count, "overall_stats")

    return {
        "student_calculations": make_category(student_pass,  student_fail),
        "assessment_stats":     make_category(assess_pass,   assess_fail),
        "overall_stats":        make_category(overall_pass,  overall_fail),
        "advanced_metrics":     make_category(advanced_pass, advanced_fail),
        "boundary_cases":       make_category(boundary_pass, boundary_fail),
    }


# -- Post-processing: failures by location -------------------------------------

def build_failures_by_location(categories):
    result = defaultdict(lambda: defaultdict(list))
    for cat in categories.values():
        for fc in cat["failed_checks"]:
            sheet   = fc.get("sheet", "Unknown")
            row     = fc.get("row")
            entity  = fc.get("entity", "unknown")
            col_name = fc.get("column_name", fc["check"])
            actual  = fc.get("actual")
            expected = fc.get("expected")
            d       = fc.get("delta")
            delta_str = f" (delta {d})" if d is not None else ""
            row_key = f"row_{row}_{entity}" if row else f"row_unknown_{entity}"
            result[sheet][row_key].append(
                f"{col_name}: expected {expected} got {actual}{delta_str}"
            )
    # Convert defaultdicts to plain dicts for JSON serialisation
    return {sheet: dict(rows) for sheet, rows in result.items()}


# -- Post-processing: error pattern summary ------------------------------------

def _diagnose_column(col_name, sheet, col_letter, fails, gt):
    """
    Return a specific root-cause hypothesis string for a column with multiple failures,
    or None if no clear pattern is detectable.
    """
    n        = len(fails)
    actuals  = [f.get("actual")   for f in fails]
    expecteds= [f.get("expected") for f in fails]
    deltas   = [f.get("delta")    for f in fails if f.get("delta") is not None]
    ref      = f"{sheet}!{col_letter} ({col_name})"

    # --- None / missing column ------------------------------------------------
    none_count = sum(1 for a in actuals if a is None)
    if none_count == n:
        return (f"{ref}: ALL {n} values are None — column is missing from the sheet "
                f"or skill crashed before writing it")
    if none_count >= n * 0.6:
        return (f"{ref}: {none_count}/{n} values are None — skill likely crashing "
                f"partway through (divide-by-zero or missing column)")

    # --- Zero-std-dev z-score (all expected = 0.0, actual ≠ 0.0) -------------
    if col_name in ("z_score",) or sheet == "Z-Scores":
        zero_exp = [f for f in fails if f.get("expected") == 0.0]
        if len(zero_exp) >= 3:
            sample_actuals = [f.get("actual") for f in zero_exp[:3]]
            return (f"{ref}: {len(zero_exp)}/{n} failures expect z=0.0 — this assessment "
                    f"has std_dev=0 (all scores identical), skill must return 0.0 when "
                    f"std_dev=0 instead of dividing by zero. "
                    f"Sample actual values: {sample_actuals}")

    # --- Rounding issue (all deltas tiny, expected has fewer decimal places) --
    if deltas and max(deltas) < 0.11:
        exp_clean = [e for e in expecteds if e is not None]
        all_rounded = exp_clean and all(
            abs(round(float(e), 1) - float(e)) < 0.0001 for e in exp_clean
        )
        if all_rounded:
            return (f"{ref}: all {n} deltas are tiny (max {max(deltas):.4f}) and expected "
                    f"values are rounded to 1 dp — skill is not applying round(value, 1). "
                    f"Sample: actual={actuals[0]}, expected={expecteds[0]}")
        # Generic rounding with 2dp
        all_rounded_2 = exp_clean and all(
            abs(round(float(e), 2) - float(e)) < 0.0001 for e in exp_clean
        )
        if all_rounded_2:
            return (f"{ref}: all {n} deltas are tiny (max {max(deltas):.4f}) — skill "
                    f"is not rounding to 2 dp. Sample: actual={actuals[0]}, expected={expecteds[0]}")

    # --- Systematic offset (all deltas same sign, similar magnitude) ----------
    if deltas and len(deltas) >= 3:
        signed = []
        for f in fails:
            try:
                signed.append(float(f["actual"]) - float(f["expected"]))
            except (TypeError, ValueError):
                pass
        if signed:
            all_pos = all(s > 0 for s in signed)
            all_neg = all(s < 0 for s in signed)
            if all_pos or all_neg:
                avg_off = sum(signed) / len(signed)
                direction = "consistently high" if all_pos else "consistently low"
                return (f"{ref}: actual values are {direction} by ~{abs(avg_off):.4f} "
                        f"on average across all {n} rows — likely a wrong constant, "
                        f"wrong denominator, or off-by-one in the formula")

    # --- Sign flip (actual ≈ -expected) ---------------------------------------
    sign_flips = 0
    for f in fails:
        try:
            if abs(float(f["actual"]) + float(f["expected"])) < 0.01:
                sign_flips += 1
        except (TypeError, ValueError):
            pass
    if sign_flips >= n * 0.5:
        return (f"{ref}: {sign_flips}/{n} actual values appear to be the sign-opposite "
                f"of expected — OLS slope formula likely has a sign error (numerator or "
                f"denominator negated)")

    # --- Constant actual (skill writing same value to every row) --------------
    non_none = [a for a in actuals if a is not None]
    if non_none and len(set(str(a) for a in non_none)) == 1:
        return (f"{ref}: skill wrote the same value ({non_none[0]}) to all {n} rows — "
                f"likely computing a scalar once and not iterating per-student/per-assessment")

    # --- Wrong denominator for percentile (actual ≈ expected * n/(n-1)) ------
    if col_name == "percentile" and deltas:
        n_students = len(gt.get("per_student", {}))
        scale_errors = 0
        for f in fails:
            try:
                ratio = float(f["actual"]) / float(f["expected"])
                if abs(ratio - n_students / (n_students - 1)) < 0.01:
                    scale_errors += 1
            except (TypeError, ValueError, ZeroDivisionError):
                pass
        if scale_errors >= n * 0.5:
            return (f"{ref}: {scale_errors}/{n} actual values match the pattern of "
                    f"dividing by n ({n_students}) instead of (n-1) ({n_students-1}) — "
                    f"fix: use (count_below / (n-1)) * 100, not (count_below / n) * 100")

    return None


def build_error_pattern_summary(categories, gt):
    n_students    = len(gt.get("per_student", {}))
    n_assessments = len(gt.get("per_assessment", {}))

    col_errors = defaultdict(list)
    row_errors = defaultdict(lambda: {"sheet": "", "row": None, "entity": "", "fields": []})

    for cat in categories.values():
        for fc in cat["failed_checks"]:
            sheet    = fc.get("sheet", "Unknown")
            col      = fc.get("column_letter", "?")
            col_name = fc.get("column_name", "?")
            row      = fc.get("row")
            entity   = fc.get("entity", "?")
            col_errors[(sheet, col, col_name)].append(fc)
            rk = f"{sheet}__row_{row}"
            row_errors[rk]["sheet"]  = sheet
            row_errors[rk]["row"]    = row
            row_errors[rk]["entity"] = entity
            row_errors[rk]["fields"].append(col_name)

    column_level, row_level, isolated = [], [], []
    root_cause_hypotheses = []

    for (sheet, col, col_name), fails in sorted(col_errors.items(), key=lambda x: -len(x[1])):
        n     = len(fails)
        denom = n_students if sheet in ("Student Stats", "Z-Scores") else n_assessments
        summary = f"{col_name} ({sheet} col {col}): wrong in {n}/{denom} rows"
        if n >= 3:
            column_level.append(summary)
            hypothesis = _diagnose_column(col_name, sheet, col, fails, gt)
            if hypothesis:
                root_cause_hypotheses.append(hypothesis)
        elif n == 1:
            entity = fails[0].get("entity", "?")
            isolated.append(f"{col_name} ({sheet} col {col}): single error for {entity} — "
                            f"actual={fails[0].get('actual')}, expected={fails[0].get('expected')}")

    for rk, info in sorted(row_errors.items(), key=lambda x: -len(x[1]["fields"])):
        n = len(info["fields"])
        if n >= 3:
            row_level.append(
                f"{info['entity']} (row {info['row']} in {info['sheet']}): "
                f"{n} errors across {', '.join(info['fields'][:4])} — "
                f"may indicate a data-parsing or row-level formula issue"
            )

    # Cross-sheet correlations
    cross_sheet = []
    wa_errors, z_errors, grade_errors = set(), set(), set()
    for cat in categories.values():
        for fc in cat["failed_checks"]:
            parts = fc["check"].split(".")
            if len(parts) == 3 and parts[0] == "student":
                sid, field = parts[1], parts[2]
                if field == "weighted_average": wa_errors.add(sid)
                elif field == "z_score":        z_errors.add(sid)
                elif field == "letter_grade":   grade_errors.add(sid)

    if wa_errors & z_errors:
        cross_sheet.append(
            f"weighted_average errors propagate to z_score for {len(wa_errors & z_errors)} "
            f"student(s) — fix weighted_average first to resolve downstream z_score failures"
        )
    if wa_errors & grade_errors:
        cross_sheet.append(
            f"weighted_average errors causing letter_grade errors in "
            f"{len(wa_errors & grade_errors)} row(s) — same students affected in both columns"
        )
    if wa_errors:
        cross_sheet.append(
            f"weighted_average errors in {len(wa_errors)} rows will also corrupt "
            f"overall.class_weighted_average, percentile, and rank downstream"
        )

    return {
        "column_level_errors":    column_level[:8],
        "row_level_errors":       row_level[:5],
        "isolated_errors":        isolated[:6],
        "cross_sheet_errors":     cross_sheet,
        "root_cause_hypotheses":  root_cause_hypotheses,
    }


# -- Excel parser --------------------------------------------------------------

def parse_excel_output(xlsx_path):
    """
    Parse the skill-generated Excel file.
    Returns (parsed_data, location_map) where location_map maps
    each sheet to its column-letter and row-number indexes.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def build_sheet_location(ws, key_col_idx=0):
        """Return (col_map, row_map) for a sheet."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}, {}
        headers = rows[0]
        col_map = {}
        for i, h in enumerate(headers):
            if h is not None:
                col_map[norm(str(h))] = get_column_letter(i + 1)
        row_map = {}
        for r_idx, row in enumerate(rows[1:], start=2):
            if row and row[key_col_idx] is not None:
                key_raw  = str(row[key_col_idx])
                key_norm = norm(key_raw)
                row_map[key_raw]  = r_idx
                row_map[key_norm] = r_idx
        return col_map, row_map

    location_map = {}

    def sheet_to_records(sheet_name):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [norm(h) for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(c is not None for c in row)]

    # -- Student Stats ---------------------------------------------------------
    per_student = {}
    if "Student Stats" in wb.sheetnames:
        ws = wb["Student Stats"]
        col_map, row_map = build_sheet_location(ws, key_col_idx=0)
        location_map["Student Stats"] = {"columns": col_map, "rows": row_map}
        for rec in sheet_to_records("Student Stats"):
            sid = str(rec.get("student_id", rec.get("student", "")))
            per_student[sid] = {
                "simple_average":      rec.get("simple_average", rec.get("average")),
                "weighted_average":    rec.get("weighted_average", rec.get("weighted_avg")),
                "letter_grade":        rec.get("letter_grade", rec.get("grade")),
                "slope":               rec.get("slope"),
                "std_dev":             rec.get("std_dev", rec.get("standard_deviation")),
                "highest_score":       rec.get("highest_score", rec.get("max")),
                "lowest_score":        rec.get("lowest_score", rec.get("min")),
                "rank":                rec.get("rank"),
                "percentile":          rec.get("percentile"),
                "z_score":             rec.get("z_score"),
                "above_class_average": rec.get("above_class_average", rec.get("above_avg")),
            }

    # -- Assessment Stats ------------------------------------------------------
    per_assessment = {}
    assess_sheet = next((s for s in ["Assessment Stats", "Test Stats"] if s in wb.sheetnames), None)
    if assess_sheet:
        ws = wb[assess_sheet]
        col_map, row_map = build_sheet_location(ws, key_col_idx=0)
        location_map["Assessment Stats"] = {"columns": col_map, "rows": row_map}
        for rec in sheet_to_records(assess_sheet):
            aname = str(rec.get("assessment", rec.get("assessment_name",
                        rec.get("test", rec.get("test_name", "")))))
            per_assessment[aname] = {
                "class_average":        rec.get("class_average", rec.get("average")),
                "std_dev":              rec.get("std_dev", rec.get("standard_deviation")),
                "highest_score":        rec.get("highest_score", rec.get("max")),
                "lowest_score":         rec.get("lowest_score", rec.get("min")),
                "per_student_z_scores": {},
            }

    # -- Z-Scores --------------------------------------------------------------
    if "Z-Scores" in wb.sheetnames:
        ws = wb["Z-Scores"]
        col_map, row_map = build_sheet_location(ws, key_col_idx=0)
        location_map["Z-Scores"] = {"columns": col_map, "rows": row_map}
        z_rows = list(ws.iter_rows(values_only=True))
        if z_rows:
            z_headers = [norm(h) for h in z_rows[0]]
            for row in z_rows[1:]:
                if not any(c is not None for c in row):
                    continue
                rec = dict(zip(z_headers, row))
                sid = str(rec.get("student_id", rec.get("student", "")))
                for col_name in z_headers:
                    if col_name in ("student_id", "student"):
                        continue
                    z_val = rec.get(col_name)
                    if col_name not in per_assessment:
                        per_assessment[col_name] = {
                            "class_average": None, "std_dev": None,
                            "highest_score": None, "lowest_score": None,
                            "per_student_z_scores": {},
                        }
                    per_assessment[col_name]["per_student_z_scores"][sid] = z_val

    # -- Overall ---------------------------------------------------------------
    overall_raw = {}
    if "Overall" in wb.sheetnames:
        ws = wb["Overall"]
        # Build row_map keyed by normalised key-column value
        overall_row_map = {}
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and row[0] is not None:
                k = norm(str(row[0]))
                overall_raw[k] = row[1] if len(row) > 1 else None
                overall_row_map[k]        = r_idx
                overall_row_map[str(row[0])] = r_idx
        location_map["Overall"] = {
            "columns": {"key": "A", "value": "B"},
            "rows":    overall_row_map,
        }

    grade_dist = {}
    for g in ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]:
        key = norm(f"grade_{g}")
        if key in overall_raw:
            grade_dist[g] = overall_raw[key]

    overall = {
        "class_weighted_average":  overall_raw.get("class_weighted_average",
                                   overall_raw.get("class_average")),
        "class_median":            overall_raw.get("class_median"),
        "grade_distribution":      grade_dist,
        "most_improved_student":   overall_raw.get("most_improved_student"),
        "most_consistent_student": overall_raw.get("most_consistent_student"),
    }

    parsed = {"per_student": per_student, "per_assessment": per_assessment, "overall": overall}
    return parsed, location_map


# -- Summary Excel generator ---------------------------------------------------

def generate_summary_excel(version, output_dir, gt_dir_path, logs_dir_path):
    """
    Write /logs/eval_results/<version>/benchmark_summary.xlsx with one sheet
    per test input.  Each row = one student.  Columns = raw scores + expected/
    actual pairs for every metric + Overall_Pass.

    Row fill:  green  = all per-student checks pass
               red    = at least one check fails
    Cell fill: the specific "actual" cells that are wrong get a darker red
               so the master LLM can spot individual errors at a glance.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    GREEN      = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_ROW    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    RED_CELL   = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    BOLD       = Font(bold=True)
    BOLD_RED   = Font(bold=True, color="9C0006")
    CENTER     = Alignment(horizontal="center")

    # (display header, gt_key, parsed_key)
    METRIC_COLS = [
        ("Weighted_Avg", "weighted_average", "weighted_average"),
        ("Simple_Avg",   "simple_average",   "simple_average"),
        ("Grade",        "letter_grade",      "letter_grade"),
        ("Slope",        "slope",             "slope"),
        ("StdDev",       "std_dev",           "std_dev"),
        ("Percentile",   "percentile",        "percentile"),
        ("ZScore",       "z_score",           "z_score"),
        ("Rank",         "rank",              "rank"),
    ]

    headers = (
        ["Student_ID",
         "Quiz_1", "Quiz_2", "Quiz_3", "Quiz_4",
         "Test_1", "Test_2", "Test_3", "Test_4", "Test_5", "Test_6"]
        + [f"{m}_Expected" for m, _, _ in METRIC_COLS]
        + [f"{m}_Actual"   for m, _, _ in METRIC_COLS]
        + ["Overall_Pass"]
    )

    # Column index of first "actual" column (0-based within headers list)
    first_actual_col = 11 + len(METRIC_COLS)   # after ID + 10 scores + expected cols

    wb = Workbook()
    wb.remove(wb.active)

    for csv_path in sorted(INPUTS_DIR.glob("*.csv")):
        gt_path   = gt_dir_path   / (csv_path.stem + ".json")
        xlsx_path = output_dir    / (csv_path.stem + "_output.xlsx")
        if not gt_path.exists() or not xlsx_path.exists():
            continue

        gt = json.loads(gt_path.read_text())
        try:
            parsed, _ = parse_excel_output(xlsx_path)
        except Exception as e:
            print(f"  [summary] Could not parse {xlsx_path.name}: {e}")
            continue

        # Sheet name: "01 Clean", "03 Ties", etc.
        sheet_title = csv_path.stem.replace("_", " ").title()[:31]
        ws = wb.create_sheet(title=sheet_title)

        # Header row
        for c_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=hdr)
            cell.font      = BOLD
            cell.alignment = CENTER
        ws.freeze_panes = "A2"

        # Data rows
        for r_idx, (sid, gt_s) in enumerate(gt["per_student"].items(), 2):
            p_s    = parsed["per_student"].get(sid, {})
            scores = gt_s["scores"]   # [q1..q4, t1..t6]

            expected_vals = [gt_s.get(gt_k)      for _, gt_k, _ in METRIC_COLS]
            actual_vals   = [p_s.get(parsed_k)   for _, _, parsed_k in METRIC_COLS]

            # Which actual values fail?
            fails = [
                not _close_enough(a, e)
                for a, e in zip(actual_vals, expected_vals)
            ]
            row_pass = not any(fails)
            row_fill = GREEN if row_pass else RED_ROW

            row_data = (
                [sid] + scores
                + expected_vals
                + actual_vals
                + ["PASS" if row_pass else "FAIL"]
            )

            for c_idx, value in enumerate(row_data, 1):
                cell       = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.fill  = row_fill
                cell.alignment = CENTER

                # Darken specific failing "actual" cells
                actual_offset = c_idx - 1 - (11 + len(METRIC_COLS))  # 0-based
                if 0 <= actual_offset < len(fails) and fails[actual_offset]:
                    cell.fill = RED_CELL
                    cell.font = BOLD_RED

        # Autofit: measure max content width per column
        for c_idx, hdr in enumerate(headers, 1):
            col_letter = get_column_letter(c_idx)
            max_w = len(hdr)
            for row in ws.iter_rows(min_row=2, min_col=c_idx, max_col=c_idx):
                for cell in row:
                    if cell.value is not None:
                        max_w = max(max_w, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_w + 3, 30)

    out_path = logs_dir_path / "benchmark_summary.xlsx"
    wb.save(out_path)
    return out_path


# -- Skill runner --------------------------------------------------------------

def run_skill(skill_path, csv_path, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / (csv_path.stem + "_output.xlsx")

    prompt = f"""You are executing the student grades skill defined in {skill_path}.

Input CSV:        {csv_path}
Output Excel:     {xlsx_path}

Execute the skill exactly as specified. As you work, produce a detailed execution
trace by explicitly logging each of the following decision points:

SLOPE FORMULA
- State the exact formula you are using for slope (OLS, numpy polyfit, or other)
- Confirm the x-axis values you are using (e.g. [1,2,3,4,5,6,7,8,9,10])
- Show the computed slope for at least the first student as a sanity check

LETTER GRADE BOUNDARIES
- List the exact grade boundary thresholds you are applying (including plus/minus)
- For each student, state which boundary their weighted_average falls into and the grade assigned
- Flag any student whose weighted_average is within 0.5 points of a boundary

RANK TIE-BREAKING
- Describe the tie-breaking strategy you are using (dense rank, competition rank, etc.)
- Identify any tied students and explicitly state the rank assigned to each
- If no ties exist in this dataset, state that explicitly

STANDARD DEVIATION
- Confirm whether you are using sample std dev (ddof=1) or population (ddof=0)
- Show the computed std dev for at least the first student

WEIGHTED AVERAGE
- State the weights: quizzes = 5% each (4 quizzes = 20% total), tests = 80/6% each (6 tests = 80% total)
- Show the computed weighted_average for at least the first student

PERCENTILE AND Z-SCORE
- Explain the percentile formula: (count strictly below / (n-1)) * 100
- Explain z_score formula: (student_wa - class_wa_mean) / class_wa_sample_std

INTERPRETATION CHOICES
- Note any point where the skill specification was ambiguous and state your choice
- Note any edge case in the data (e.g. all-zero scores, perfect scores, missing values)
  and how you handled it

After completing the Excel output, print a one-line summary confirming the output
file was written and the sheet names used.
"""

    trace = ""
    try:
        result = subprocess.run(
            ["claude", "-p", prompt, "--allowedTools", "Bash,Write,Read"],
            capture_output=True, text=True, timeout=300
        )
        trace = result.stdout + result.stderr
        if not xlsx_path.exists():
            return None, trace
    except Exception as e:
        return None, str(e)

    return xlsx_path, trace


# -- Version diff --------------------------------------------------------------

def load_previous_report(test_stem, current_version):
    candidates = []
    for version_dir in LOGS_BASE.iterdir():
        if version_dir.is_dir() and version_dir.name != current_version:
            candidates.extend(version_dir.glob(f"{test_stem}_*.json"))
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for path in candidates:
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            continue
    return None


def diff_reports(current, previous):
    def flatten_checks(report):
        result = {}
        for cat in report.get("categories", {}).values():
            for c in cat.get("passed_checks", []):
                result[c["check"]] = True
            for c in cat.get("failed_checks", []):
                result[c["check"]] = False
        return result

    cur_checks  = flatten_checks(current)
    prev_checks = flatten_checks(previous)
    improved, regressed = [], []
    unchanged_pass = unchanged_fail = 0
    for name, cur_passed in cur_checks.items():
        prev_passed = prev_checks.get(name)
        if prev_passed is None:
            continue
        if not prev_passed and cur_passed:   improved.append(name)
        elif prev_passed and not cur_passed: regressed.append(name)
        elif cur_passed:  unchanged_pass += 1
        else:             unchanged_fail += 1
    return {
        "previous_version": previous.get("skill_version", "unknown"),
        "improved":         improved,
        "regressed":        regressed,
        "unchanged_pass":   unchanged_pass,
        "unchanged_fail":   unchanged_fail,
    }


# -- Console summary -----------------------------------------------------------

CATEGORY_LABELS = {
    "student_calculations": "Student Calculations",
    "assessment_stats":     "Assessment Stats    ",
    "overall_stats":        "Overall Stats       ",
    "advanced_metrics":     "Advanced Metrics    ",
    "boundary_cases":       "Boundary Cases      ",
}

def _top_column_failures(failed_checks, top_n=3):
    """Return top N (sheet, col, col_name, entities, min_row, max_row) by error count."""
    col_data = defaultdict(lambda: {"entities": [], "rows": []})
    for fc in failed_checks:
        sheet    = fc.get("sheet", "?")
        col      = fc.get("column_letter", "?")
        col_name = fc.get("column_name", fc.get("check", "?"))
        entity   = fc.get("entity", "?")
        row      = fc.get("row")
        key = (sheet, col, col_name)
        col_data[key]["entities"].append(entity)
        if row:
            col_data[key]["rows"].append(row)
    sorted_cols = sorted(col_data.items(), key=lambda x: -len(x[1]["entities"]))[:top_n]
    result = []
    for (sheet, col, col_name), data in sorted_cols:
        rows = sorted(data["rows"])
        rng  = f"{col}{rows[0]}:{col}{rows[-1]}" if len(rows) > 1 else (f"{col}{rows[0]}" if rows else col)
        result.append((sheet, col, col_name, data["entities"], rng))
    return result

def print_summary_table(test_name, categories):
    print(f"\n  {test_name}")
    total_p = total_t = 0
    for key, label in CATEGORY_LABELS.items():
        cat = categories.get(key, {})
        p, t  = cat.get("passed", 0), cat.get("total", 0)
        total_p += p
        total_t += t
        s   = cat.get("score", 0.0)
        sym = "OK" if p == t else "!!"
        print(f"    {label}  {fraction(p,t):>10}  ({s:.2f})  {sym}")
        if p < t:
            for sheet, col, col_name, entities, rng in _top_column_failures(cat["failed_checks"]):
                n = len(entities)
                print(f"      -> {sheet}!{rng}  {col_name}: wrong in {n} row(s)")
    print(f"    {'--' * 26}")
    overall_s = score(total_p, total_t)
    sym = "OK" if total_p == total_t else "!!"
    print(f"    {'Overall':22}  {fraction(total_p, total_t):>10}  ({overall_s:.2f})  {sym}")


def print_diff(diff):
    prev = diff["previous_version"]
    print(f"\n  Diff vs {prev}:")
    if diff["improved"]:
        print(f"    [+] Improved  ({len(diff['improved'])}): "
              + ", ".join(diff["improved"][:5]) + (" ..." if len(diff["improved"]) > 5 else ""))
    if diff["regressed"]:
        print(f"    [-] Regressed ({len(diff['regressed'])}): "
              + ", ".join(diff["regressed"][:5]) + (" ..." if len(diff["regressed"]) > 5 else ""))
    print(f"    Unchanged: {diff['unchanged_pass']} passing, {diff['unchanged_fail']} failing")


# -- Main ----------------------------------------------------------------------

def run_benchmark(version):
    skill_path = SKILL_DIR / f"skill_{version}.md"
    if not skill_path.exists():
        print(f"Skill file not found: {skill_path}")
        sys.exit(1)

    output_dir = ROOT / "outputs" / version
    logs_dir   = LOGS_BASE / version
    logs_dir.mkdir(parents=True, exist_ok=True)
    csv_files  = sorted(INPUTS_DIR.glob("*.csv"))
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    all_summaries = []

    for csv_path in csv_files:
        gt_path = GT_DIR / (csv_path.stem + ".json")
        if not gt_path.exists():
            print(f"No ground truth for {csv_path.name}, skipping.")
            continue

        with open(gt_path) as f:
            gt = json.load(f)

        print(f"\nRunning: {csv_path.name} ...")
        xlsx_path, trace = run_skill(skill_path, csv_path, output_dir)

        if xlsx_path is None:
            categories = {k: make_category([], []) for k in CATEGORY_LABELS}
            report = {
                "test_input":      csv_path.name,
                "skill_version":   version,
                "timestamp":       timestamp,
                "execution_trace": trace,
                "error":           "Skill did not produce Excel output.",
                "overall":         {"fraction": "0/0", "score": 0.0, "passed": 0, "total": 0},
                "categories":      categories,
                "failures_by_location":  {},
                "error_pattern_summary": {},
            }
        else:
            try:
                parsed, location_map = parse_excel_output(xlsx_path)
                categories           = compare(parsed, gt, location_map)
                total_p = sum(c["passed"] for c in categories.values())
                total_t = sum(c["total"]  for c in categories.values())
                report = {
                    "test_input":      csv_path.name,
                    "skill_version":   version,
                    "timestamp":       timestamp,
                    "execution_trace": trace,
                    "overall": {
                        "fraction": fraction(total_p, total_t),
                        "score":    score(total_p, total_t),
                        "passed":   total_p,
                        "total":    total_t,
                    },
                    "categories":             categories,
                    "failures_by_location":   build_failures_by_location(categories),
                    "error_pattern_summary":  build_error_pattern_summary(categories, gt),
                }
            except Exception as e:
                import traceback
                categories = {k: make_category([], []) for k in CATEGORY_LABELS}
                report = {
                    "test_input":      csv_path.name,
                    "skill_version":   version,
                    "timestamp":       timestamp,
                    "execution_trace": trace,
                    "error":           f"Parse/compare error: {e}\n{traceback.format_exc()}",
                    "overall":         {"fraction": "0/0", "score": 0.0, "passed": 0, "total": 0},
                    "categories":      categories,
                    "failures_by_location":  {},
                    "error_pattern_summary": {},
                }

        prev_report = load_previous_report(csv_path.stem, version)
        if prev_report:
            report["version_diff"] = diff_reports(report, prev_report)

        out_file = logs_dir / f"{csv_path.stem}_{version}_{timestamp}.json"
        with open(out_file, "w") as f:
            json.dump(report, f, indent=2, default=str)

        print_summary_table(csv_path.name, report["categories"])
        hypotheses = report.get("error_pattern_summary", {}).get("root_cause_hypotheses", [])
        if hypotheses:
            print(f"\n  Root cause hypotheses:")
            for h in hypotheses:
                print(f"    !! {h}")
        if prev_report and "version_diff" in report:
            print_diff(report["version_diff"])

        all_summaries.append(report)

    if all_summaries:
        print(f"\n{'=' * 52}")
        print(f"  BENCHMARK COMPLETE  --  skill {version}")
        print(f"{'=' * 52}")
        agg_p = sum(r["overall"]["passed"] for r in all_summaries)
        agg_t = sum(r["overall"]["total"]  for r in all_summaries)
        print(f"  Total: {fraction(agg_p, agg_t)}  ({score(agg_p, agg_t):.2f})")

        summary_path = generate_summary_excel(version, output_dir, GT_DIR, logs_dir)
        print(f"  Summary Excel: {summary_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--version", required=True, help="Skill version, e.g. v0")
    args = parser.parse_args()
    run_benchmark(args.version)
